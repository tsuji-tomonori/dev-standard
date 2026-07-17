from __future__ import annotations

import hashlib
import os
from copy import copy
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation

BOOK = Path("checklist.xlsx")
SWEBOK = Path(os.environ.get("SWEBOK_PDF", ".workspace/swebok-v4a.pdf"))
EXPECTED_SWEBOK_SHA256 = "b3cb8028fecb9607f757504c861947fa3bf423087ea8bf08c58020f0ba3596dc"
UPDATED = date(2026, 7, 18)

REVIEW_SHEETS = [
    "01_要件定義",
    "02_アーキテクチャ",
    "03_詳細設計",
    "04_実装",
    "05_テスト",
    "06_運用",
    "07_保守",
    "08_構成管理",
    "09_マネジメント",
    "10_プロセス",
    "12_品質保証",
    "13_セキュリティ",
    "14_クラウド",
    "15_AWS",
    "16_GoogleCloud",
    "17_Azure",
    "18_OCI",
    "19_AI",
]

PROFILE_BY_SHEET = {
    **{name: "CORE" for name in REVIEW_SHEETS[:12]},
    "14_クラウド": "CLOUD-COMMON",
    "15_AWS": "AWS-DELTA",
    "16_GoogleCloud": "GCP-DELTA",
    "17_Azure": "AZURE-DELTA",
    "18_OCI": "OCI-DELTA",
    "19_AI": "AI-CONDITIONAL",
}

EXT_HEADERS = [
    "適用プロファイル",
    "適用条件",
    "出典ID",
    "重複統制グループ",
    "案件重要度",
    "適用判定",
    "判定",
    "証跡・参照箇所",
    "指摘・Issue ID",
    "対応方針",
    "例外・リスク受容承認者",
    "是正期限",
    "レビュアー",
    "レビュー日",
    "再確認結果",
]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def find_row(ws, item_id: str) -> int:
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == item_id:
            return row
    raise KeyError(f"{ws.title}: {item_id}")


def normalize_swebok_references(wb) -> None:
    for sheet in REVIEW_SHEETS:
        ws = wb[sheet]
        for row in range(2, ws.max_row + 1):
            reference = ws.cell(row, 8).value
            if isinstance(reference, str):
                ws.cell(row, 8).value = reference.replace("swebok-v4.pdf", "swebok-v4a.pdf")
    summary = wb["サマリ"]
    for row in range(5, 5 + len(REVIEW_SHEETS)):
        target = summary.cell(row, 3).value
        if isinstance(target, str):
            summary.cell(row, 3).value = target.replace("swebok-v4.pdf", "swebok-v4a.pdf")


def set_item(ws, item_id: str, *, check: str, criterion: str, intent: str, reference: str | None = None) -> None:
    row = find_row(ws, item_id)
    ws.cell(row, 5).value = check
    ws.cell(row, 6).value = criterion
    ws.cell(row, 7).value = intent
    if reference is not None:
        ws.cell(row, 8).value = reference
    ws.cell(row, 5).comment = Comment(intent, "Checklist maintainer")


def insert_items(ws, parent_id: str, items: list[dict[str, str]]) -> None:
    existing = {ws.cell(r, 1).value: r for r in range(2, ws.max_row + 1)}
    missing = [item for item in items if item["id"] not in existing]
    if missing:
        parent_row = find_row(ws, parent_id)
        ws.insert_rows(parent_row + 1, len(missing))
        for offset, item in enumerate(missing, 1):
            target = parent_row + offset
            for col in range(1, 9):
                source = ws.cell(parent_row, col)
                cell = ws.cell(target, col)
                if source.has_style:
                    cell._style = copy(source._style)
                cell.number_format = source.number_format
                cell.alignment = copy(source.alignment)
                cell.protection = copy(source.protection)
            ws.row_dimensions[target].height = ws.row_dimensions[parent_row].height
            values = [
                item["id"],
                item.get("section", ws.cell(parent_row, 2).value),
                item.get("category", ws.cell(parent_row, 3).value),
                item.get("severity", ws.cell(parent_row, 4).value),
                item["check"],
                item["criterion"],
                item["intent"],
                item.get("reference", ws.cell(parent_row, 8).value),
            ]
            for col, value in enumerate(values, 1):
                ws.cell(target, col).value = value
            ws.cell(target, 5).comment = Comment(item["intent"], "Checklist maintainer")


def apply_content_fixes(wb) -> None:
    sec = wb["13_セキュリティ"]
    set_item(
        sec,
        "SEC-005",
        check="信頼境界を越える各データフローについて、脅威の有無と分析根拠が記録されている",
        criterion="境界を越える各フローにSTRIDE等による脅威が対応付くか、脅威なしの場合は資産・攻撃面・既存統制に基づく非該当理由が記録されている",
        intent="機械的に脅威を1件以上作るのではなく、全フローを分析した証跡と、該当なし判断の説明可能性を確保する。",
    )
    set_item(
        sec,
        "SEC-012",
        check="脅威モデルに必要な攻撃経路を記述し、機微度に応じて文書を分類・アクセス制御している",
        criterion="防御設計と検証に必要な攻撃経路・前提・影響が記載され、詳細度に応じた情報分類、閲覧権限、共有・保管ルールが適用されている",
        intent="攻撃経路を削ると設計・検証に使えない。必要な技術情報を残したうえで、文書自体を機密資産として保護する。",
    )

    cloud = wb["14_クラウド"]
    set_item(
        cloud,
        "CLD-001",
        check="SLO・許容停止時間・障害ドメイン分析に基づき、必要な階層へゾーン冗長または同等の可用性対策を適用している",
        criterion="対象階層ごとに可用性目標、単一障害点、採用した冗長化方式（非採用理由を含む）が対応付けられ、要求される障害シナリオの試験結果が目標を満たす",
        intent="複数AZを一律の唯一解にせず、事業要求と障害モデルに比例した冗長化を選び、過少対策と過剰投資を避ける。",
    )
    set_item(
        cloud,
        "CLD-004",
        check="事業影響、SLO、RTO/RPO、許容コストに基づき、必要なDR方式と対象障害範囲を選定している",
        criterion="リージョン障害を含む対象シナリオごとに、バックアップ／復元、同一リージョン冗長、クロスリージョン複製等の採否と根拠が記録され、RTO/RPO・コスト制約と整合する",
        intent="クロスリージョンDRを一律要求せず、復旧目標と事業影響から必要な方式を選ぶ。",
    )
    insert_items(
        cloud,
        "CLD-004",
        [
            {
                "id": "CLD-065",
                "check": "採用したDR方式について、復旧・切替手順が実装され、必要なデータ複製と依存関係が構成されている",
                "criterion": "DR設計に対応する構成、権限、データ複製、DNS・接続先切替、復旧手順が環境上で確認できる",
                "intent": "選定だけで終わらせず、実際に復旧可能な構成と手順へ落とし込む。",
            },
            {
                "id": "CLD-066",
                "check": "DR切替・復旧テストを定期実施し、実測RTO/RPOと改善事項を記録している",
                "criterion": "対象障害シナリオの演習記録、実測復旧時間・データ損失量、失敗事項、是正状況があり、承認済みRTO/RPOを満たす",
                "intent": "未検証のDR計画が災害時に機能しないリスクを、実測と改善ループで抑える。",
            },
        ],
    )
    set_item(
        cloud,
        "CLD-011",
        check="スケール・フェイルオーバー対象の状態管理方式が、可用性・整合性・性能要件に基づき選定されている",
        criterion="ステートレス化、外部状態ストア、状態レプリケーション、セッション固定等の採否と根拠があり、インスタンス追加・削除・再起動時の要求を満たす試験結果がある",
        intent="ステートレス化を唯一解にせず、ワークロード特性に合う状態分割・複製方式も許容しつつ、拡張性と復旧性を検証する。",
    )

    ai = wb["19_AI"]
    set_item(
        ai,
        "AI-004",
        check="学習・チューニングデータのスキーマ、欠損、異常値を自動検証し、不合格時にパイプラインを停止する",
        criterion="データ契約と検証ルールが実装され、違反データを用いた学習が開始されないテスト記録がある",
        intent="不良入力による無駄な学習と劣化モデルの生成を、処理の入口で防ぐ。",
    )
    insert_items(
        ai,
        "AI-004",
        [
            {"id": "AI-052", "check": "学習・チューニングデータの分布と対象母集団に対する代表性を分析している", "criterion": "主要属性・サブグループ別の分布、欠落、偏りを示すEDA結果と、許容可否の判断がある", "intent": "平均的な品質だけでは見えない母集団の欠落や偏りを検出する。"},
            {"id": "AI-053", "check": "学習・チューニングデータの前処理手順を定義し、再現可能に実行している", "criterion": "ノイズ除去、重複排除、正規化等の適用条件と実装がバージョン管理され、同じ入力から同じデータセットを再生成できる", "intent": "属人的な手作業によるデータ差異と再現不能を防ぐ。"},
        ],
    )
    set_item(
        ai,
        "AI-007",
        check="個人データの収集・利用について、適用法令と組織ポリシーに基づく適法根拠（必要な場合は同意）を確認している",
        criterion="目的、データ主体、処理活動、適法根拠、通知・同意要否、所管部門の確認結果が記録されている",
        intent="明示同意を全案件へ一律要求せず、適用法令と処理目的に即した合法性・透明性を確認する。",
    )
    insert_items(
        ai,
        "AI-007",
        [
            {"id": "AI-077", "check": "AIで収集・利用する個人データを目的達成に必要な最小限へ限定している", "criterion": "データ項目ごとに利用目的と必要性が対応付けられ、不要項目が収集・学習対象から除外されている", "intent": "不要な個人データの蓄積による漏えい・規制リスクを減らす。"},
            {"id": "AI-078", "check": "機微性と再識別リスクに応じて匿名化・仮名化・合成データを適用している", "criterion": "変換方式、再識別リスク評価、鍵・対応表の分離、残余リスクが記録されている", "intent": "データ有用性を保ちながら個人への結び付きと漏えい影響を抑える。"},
            {"id": "AI-079", "check": "AIデータの保持期間と期限到来時の削除を実装している", "criterion": "データ種別ごとの保持期間、例外、削除ジョブ、削除結果の証跡がある", "intent": "目的終了後もデータが残り続けるリスクを防ぐ。"},
            {"id": "AI-080", "check": "適用される削除請求・忘れられる権利を、学習・検索インデックス・ログまで追跡して実行できる", "criterion": "請求受付から対象特定、削除・再構築、完了通知までの手順とテスト記録がある", "intent": "派生データや検索索引に個人データが残る見落としを防ぐ。"},
        ],
    )
    set_item(
        ai,
        "AI-009",
        check="プロンプト設計、RAG、ファインチューニング、モデル変更等の候補をユースケース別に比較し、最適なカスタマイズ方式を選定している",
        criterion="品質、データ、レイテンシ、セキュリティ、運用負荷、コストの比較結果と選定理由があり、採用方式の効果が測定されている",
        intent="固定順序を前提にせず、ユースケース制約に最も合う方式または組合せを選ぶ。",
    )
    set_item(
        ai,
        "AI-011",
        check="代表ユースケースと失敗モードを含む評価データセットを定義・版管理している",
        criterion="評価対象の母集団、主要ユースケース、境界・攻撃ケースをカバーするデータセットの版と由来が記録されている",
        intent="評価データの偏りやすり替わりにより、見かけ上だけ品質が上がることを防ぐ。",
    )
    insert_items(
        ai,
        "AI-011",
        [
            {"id": "AI-054", "check": "AIリリース判定に用いる品質・安全性・根拠性・公平性等の指標と合格閾値を事前定義している", "criterion": "ユースケースに必要な各指標、集計単位、信頼水準、合格閾値、例外承認者が文書化されている", "intent": "主観的な『良くなった』ではなく、再現可能な合否判断を可能にする。"},
            {"id": "AI-055", "check": "モデル・プロンプト変更時の評価を自動実行し、同一条件で再現できる", "criterion": "評価コード、依存関係、モデル・プロンプト版、実行条件、結果が保存され、同じ版で再実行できる", "intent": "変更ごとの評価漏れと環境差による比較不能を防ぐ。"},
        ],
    )
    set_item(
        ai,
        "AI-014",
        check="モデルへ渡すユーザー入力・外部コンテンツを検証し、危険な形式・命令・データを無害化または拒否している",
        criterion="入力面ごとのスキーマ、長さ、文字種、危険パターン、外部コンテンツの信頼度に応じた検証・拒否・サニタイズが実装されている",
        intent="未検証入力が推論コンテキストへ直接入り、命令乗っ取りや情報漏えいを起こすことを防ぐ。",
    )
    insert_items(
        ai,
        "AI-014",
        [
            {"id": "AI-056", "check": "システム指示、ユーザー入力、検索コンテンツ、ツール出力を構造的に分離している", "criterion": "各入力の役割と信頼境界がテンプレート・メッセージ構造・権限制御で分離され、外部データが上位指示として解釈されない", "intent": "データと命令の混同による間接プロンプトインジェクションを抑える。"},
            {"id": "AI-057", "check": "意図分類またはインジェクション検知により、許可外の目的・命令を検出している", "criterion": "許可意図と拒否条件が定義され、既知および変形攻撃に対する検知・拒否結果が記録されている", "intent": "表層的な文字列検査だけでは見逃す操作意図を検出する。"},
            {"id": "AI-058", "check": "モデル出力を用途別ポリシーで検証し、機密情報・危険内容・許可外形式を遮断している", "criterion": "出力先ごとのスキーマ、機密情報、コンテンツ安全性、許可操作の検証があり、違反出力が利用者やツールへ渡らない", "intent": "モデルが生成した不正・機密内容の後段流出を防ぐ。"},
            {"id": "AI-059", "check": "プロンプトインジェクションとジェイルブレイクの攻撃テストを継続実施している", "criterion": "直接・間接・多段・難読化攻撃を含むテストセット、実行頻度、結果、残余リスクが記録されている", "intent": "多層防御が実際の攻撃変種に耐えるかを継続確認する。"},
        ],
    )
    set_item(
        ai,
        "AI-030",
        check="モデル、プロンプト、オーケストレーションコード、ライブラリ、コンテナイメージを一意に版管理している",
        criterion="本番挙動を構成する全AI資産の不変な版識別子と変更履歴を取得できる",
        intent="暗黙更新や構成要素の組合せ不明による再現不能を防ぐ。",
    )
    insert_items(
        ai,
        "AI-030",
        [
            {"id": "AI-060", "check": "モデルカタログにモデルID、出所、用途、カスタマイズ履歴、評価・承認結果を維持している", "criterion": "稼働・候補・廃止モデルを横断するカタログがあり、各モデルの責任者と根拠資料へ遡れる", "intent": "モデルの由来と承認状態が不明なまま利用されることを防ぐ。"},
            {"id": "AI-061", "check": "評価環境で合格したAI資産の同一バージョン組合せを本番へ昇格している", "criterion": "評価結果の資産版と本番デプロイの版が一致し、昇格記録と差異検知がある", "intent": "テスト対象と本番実体の不一致により評価が無効になることを防ぐ。"},
        ],
    )
    set_item(
        ai,
        "AI-041",
        check="AI依存コンポーネントの信頼性目標を整合させ、要求に応じた冗長化・縮退方式を選定している",
        criterion="モデル、検索、オーケストレータ、ツールごとのSLOと依存関係があり、単一障害点への対策（再試行、代替、ゾーン／リージョン冗長等）の採否と根拠が記録されている",
        intent="マルチゾーン／リージョンを一律要求せず、AI機能のSLO・RTO/RPO・コスト制約に応じた対策を選ぶ。",
    )
    set_item(
        ai,
        "AI-043",
        check="AIワークロードのコストドライバーと単位コストをモデル化している",
        criterion="データ量、学習時間、クエリ、トークン、検索、ツール呼出し等の費用をユースケース・利用単位へ配賦した見積りと実績がある",
        intent="利用量増加時の費用構造と採算分岐を把握し、最適化判断の基準を作る。",
    )
    insert_items(
        ai,
        "AI-043",
        [
            {"id": "AI-062", "check": "タスク品質を満たす範囲でモデルをライトサイジングしている", "criterion": "候補モデルの品質、レイテンシ、単価比較と採用理由がタスク別に記録されている", "intent": "必要以上に大きいモデルの常用による費用・遅延を防ぐ。"},
            {"id": "AI-063", "check": "プロンプトと応答のトークン長を品質要件の範囲で最適化している", "criterion": "不要コンテキスト、重複指示、最大出力長の見直し前後で品質とトークン量を測定している", "intent": "品質に寄与しないトークン消費を抑える。"},
            {"id": "AI-064", "check": "意味・セキュリティ境界に適合する箇所へキャッシュを適用している", "criterion": "キャッシュ対象、キー、テナント境界、TTL、無効化条件と費用削減効果が記録されている", "intent": "重複推論・検索を減らしつつ、他利用者の応答混入を防ぐ。"},
            {"id": "AI-065", "check": "AI利用コストを部門・製品・テナント等の責任単位へ配賦している", "criterion": "タグ、利用者ID、プロジェクトID等で費用を追跡でき、責任者へ定期報告されている", "intent": "コスト責任が不明な共有基盤で利用が無制限に増えることを防ぐ。"},
            {"id": "AI-066", "check": "AI費用の予算・異常検知アラートと超過時の対応を設定している", "criterion": "予算閾値、単位コスト急増、利用量異常の通知先と抑制・承認手順がテストされている", "intent": "設定ミスや攻撃による急激な費用増加を早期に止める。"},
        ],
    )

    # 2025/2026 AWS lenses: distinct controls not present as atomic checks in the original catalog.
    insert_items(
        ai,
        "AI-051",
        [
            {"id": "AI-067", "category": "責任あるAI", "severity": "High", "check": "AIユースケースの上流・下流ステークホルダーと影響を受ける集団を特定している", "criterion": "データ提供者、開発・運用者、直接利用者、判断対象者、第三者への便益・潜在的害がユースケース単位で記録されている", "intent": "直接利用者だけを見て、判断対象者や下流利用者への影響を見落とすことを防ぐ。", "reference": "AWS Responsible AI Lens > Use case > Identify use case stakeholders (RAIUC02)"},
            {"id": "AI-068", "category": "責任あるAI", "severity": "High", "check": "責任あるAIを含むリリース基準を、定量評価・閾値・信頼水準を持つ二値判定として定義している", "criterion": "安全性、公平性、堅牢性、説明可能性等の該当基準に測定方法、閾値、必要信頼水準、合否が定義されている", "intent": "抽象的な原則を、リリース時に実際に判定できる統制へ変換する。", "reference": "AWS Responsible AI Lens > Release criteria"},
            {"id": "AI-069", "category": "データ・グラウンディング", "severity": "Medium", "check": "各AIデータセットの特性・由来・制約をデータシートとして管理している", "criterion": "目的、収集方法、母集団、ライセンス、機微性、既知の偏り、推奨・禁止用途、版がデータセットごとに記録されている", "intent": "出所や制約が不明なデータの再利用により、品質・権利・公平性問題が再発することを防ぐ。", "reference": "AWS Responsible AI Lens > Data preparation > RAIDP04-BP05 Document dataset characteristics using a datasheet / AWS Machine Learning Lens"},
            {"id": "AI-070", "category": "責任あるAI", "severity": "Medium", "check": "AI機能・モデルの廃止時に、ステークホルダーへの義務と残存データ・依存先を処理する", "criterion": "通知、代替手段、データ削除・保持、API・モデル停止、監査記録、異議・問い合わせ窓口を含む廃止計画がある", "intent": "AI機能停止後もデータや自動判断への依存が残り、利用者の権利・業務が損なわれることを防ぐ。", "reference": "AWS Responsible AI Lens > Monitoring > Decommissioning (RAIMON03-BP01)"},
            {"id": "AI-071", "category": "エージェントAI", "severity": "Critical", "check": "各AIエージェントの目的、権限、利用可能ツール、禁止操作、停止条件を明示し、技術的に強制している", "criterion": "エージェント単位の権限契約があり、IAM・ポリシーエンジン・スキーマ等でプロンプト外にも強制されている", "intent": "曖昧な目的やプロンプトだけの制約により、エージェントが意図外操作を行うことを防ぐ。", "reference": "AWS Agentic AI Lens > Design principles / Security > Agent goal alignment and manipulation prevention"},
            {"id": "AI-072", "category": "エージェントAI", "severity": "High", "check": "エージェントのツール呼出し、メモリアクセス、エージェント間引継ぎをエンドツーエンドで追跡できる", "criterion": "単一の相関IDで、入力、推論版、ツール引数・結果、メモリ読書き、引継ぎ、最終結果、承認を時系列再構成できる", "intent": "複数段の自律処理で、誰が何を根拠に操作したか不明になることを防ぐ。", "reference": "AWS Agentic AI Lens > Design principles > Make every agent action observable and traceable end-to-end"},
            {"id": "AI-073", "category": "エージェントAI", "severity": "Critical", "check": "エージェントの自律性レベルと行為の影響に比例した人間承認・エスカレーションを定義している", "criterion": "観察、提案、自律実行、オーケストレーション等のレベルと、金銭・権限・外部通信・不可逆操作ごとの承認条件が実装されている", "intent": "低リスク支援と高影響の自律操作を同じ統制で扱うことを防ぐ。", "reference": "AWS Agentic AI Lens > Design principles > Pair autonomy with proportionate human oversight"},
            {"id": "AI-074", "category": "エージェントAI", "severity": "High", "check": "エージェント間・ツール間の入出力を明示的な契約とスキーマで検証している", "criterion": "メッセージ、ツール引数、成功条件、信頼度、エラー、版互換性がスキーマ化され、違反時に拒否または安全停止する", "intent": "暗黙の自然言語契約により、誤解した指示や不正データが連鎖することを防ぐ。", "reference": "AWS Agentic AI Lens > Design principles > Ground autonomous behavior in explicit contracts"},
            {"id": "AI-075", "category": "エージェントAI", "severity": "High", "check": "エージェントの反復回数、ツール呼出し、時間、費用に上限を設け、無限ループを停止できる", "criterion": "タスク別の実行予算、タイムアウト、最大ステップ、停止・補償処理、超過アラートが構成されテストされている", "intent": "確率的な計画失敗が無限反復、外部副作用、費用暴走へ拡大することを防ぐ。", "reference": "AWS Agentic AI Lens > Reliability / Cost optimization / Operational excellence"},
            {"id": "AI-076", "category": "ユースケース・ガバナンス", "severity": "Medium", "check": "従来型ML、生成AI、エージェントAIの種別に応じて適用するレビュー資料とプロファイルを選んでいる", "criterion": "ユースケース種別と、Machine Learning Lens、Generative AI Lens、Responsible AI Lens、Agentic AI Lensの適用・非適用理由が記録されている", "intent": "生成AIだけの観点で従来型MLを評価する、またはエージェント固有リスクを通常の推論APIとして扱う誤りを防ぐ。", "reference": "AWS Machine Learning Lens > Distinction from the Generative AI Lens / AWS Responsible AI Lens / AWS Agentic AI Lens"},
        ],
    )

    azure = wb["17_Azure"]
    azure_splits = {
        "AZR-203": [
            ("AZR-251", "AIゲートウェイでモデル呼出しの認証・認可を一元実施している", "未認証呼出しが拒否され、利用者・アプリ・テナント単位の認可結果が記録される"),
            ("AZR-252", "AIゲートウェイでレート制限とトークンクォータを強制している", "利用者・テナント・モデル別のTPM/RPM・トークン上限と超過時動作がテストされている"),
            ("AZR-253", "AIゲートウェイで入出力の安全性・機密情報フィルタを適用している", "全モデル経路に同じ最低限の入出力ポリシーが適用され、迂回経路がない"),
            ("AZR-254", "AIゲートウェイのモデルルーティング規則を品質・可用性・データ制約に基づき定義している", "ルーティングとフォールバック条件、禁止モデル、データ所在制約が構成として確認できる"),
            ("AZR-255", "AIゲートウェイの利用量を費用責任単位へ配賦できる", "部門・製品・テナント等の単位でトークン・要求・費用を集計できる"),
        ],
        "AZR-208": [
            ("AZR-256", "直接のモデル呼出しでなくエージェントを採用する必要性を評価している", "複数段計画・ツール利用・状態保持等の必要性と追加リスクを比較した記録がある"),
            ("AZR-257", "採用するエージェントフレームワークでエージェント・ツール・メモリの境界を定義している", "責務、入出力、権限、状態所有者がコンポーネント単位で記録されている"),
            ("AZR-258", "決定的ワークフローと探索的エージェント協調を使い分けている", "再現性・監査性が必要な処理は決定的に制御し、探索が必要な箇所だけに自律判断を限定している"),
        ],
        "AZR-220": [
            ("AZR-259", "公開済み特徴量を不変とし、変更を新バージョンとして管理している", "利用中特徴量が上書きされず、版ごとの定義・生成コード・利用モデルへ遡れる"),
            ("AZR-260", "特徴量の命名規約と所有者を定義している", "意味、型、時間基準、所有者を識別できる名称・メタデータがある"),
            ("AZR-261", "特徴量のドリフトと品質劣化を自動検知している", "欠損、分布、鮮度、異常値の監視と閾値超過時の通知がある"),
            ("AZR-262", "未使用・低品質な特徴量を廃止する手順がある", "利用関係を確認したうえで非推奨化・削除し、影響モデルを再評価する記録がある"),
        ],
        "AZR-232": [
            ("AZR-263", "データ主権要件に基づきソブリンクラウド利用の要否を評価している", "対象データ・処理・運用者と地域要件を対応付け、採否を承認している"),
            ("AZR-264", "Microsoft Purview等で対象データをカタログ化・分類し、系譜を追跡している", "ソースからAI利用先まで分類ラベルと変換履歴を確認できる"),
            ("AZR-265", "機密性要件に応じConfidential Computingとカスタマーマネージドキーを評価・適用している", "脅威モデルに基づく採否、鍵管理、実装確認がある"),
            ("AZR-266", "対象リージョンで必要なAzureサービスと機能が提供されることを事前確認している", "必要SKU・クォータ・可用性ゾーン・プレビュー制約の確認記録がある"),
        ],
        "AZR-239": [
            ("AZR-267", "推論エンドポイントをblue-green／カナリア等で安全に更新できる", "トラフィック分割、評価、停止条件、ロールバックがテストされている"),
            ("AZR-268", "検索インデックスをside-by-side構築し、エイリアス切替で安全に更新できる", "旧新インデックスの並行検証、切替、戻し手順がある"),
            ("AZR-269", "オーケストレーションコードをフィーチャーフラグまたはblue-greenで安全に更新できる", "段階公開、依存互換性確認、即時無効化が可能である"),
        ],
        "AZR-240": [
            ("AZR-270", "エージェントの意図解決・ツール選択・タスク遵守を独立指標で評価している", "各指標の評価セットと閾値があり、変更前後を比較できる"),
            ("AZR-271", "エージェント依存先をモック化し、タイムアウト・誤応答・部分障害を試験している", "主要依存先の障害注入と期待する縮退・停止結果が記録されている"),
            ("AZR-272", "エージェントをシナリオベースで評価し、モデル採点と人手レビューを組み合わせている", "代表・境界・高影響シナリオに機械評価と人手確認の記録がある"),
        ],
        "AZR-243": [
            ("AZR-273", "グラウンディング処理の機能・統合テストを実施している", "前処理、チャンキング、埋め込み、索引、検索を通した期待結果が確認できる"),
            ("AZR-274", "グラウンディングデータのスキーマ後方互換性をテストしている", "旧新生成物・利用側の互換性と移行手順が確認できる"),
            ("AZR-275", "グラウンディングデータの鮮度・品質を自動検査している", "更新遅延、欠損、重複、破損の閾値と失敗時動作がある"),
            ("AZR-276", "検索インデックスの想定負荷試験を実施している", "検索量、更新量、同時実行のピークでSLOを満たす結果がある"),
            ("AZR-277", "グラウンディング検索のアクセス制御をセキュリティテストしている", "権限の異なる利用者で非許可文書が検索・生成結果に含まれない"),
        ],
        "AZR-248": [
            ("AZR-278", "エージェント活動を改ざん耐性のある監査ログへ記録している", "ツール操作、承認、権限、結果を利用者と実行単位へ追跡できる"),
            ("AZR-279", "エージェントとツールにRBAC・最小権限を適用している", "各エージェントの職務に不要な操作が権限上拒否される"),
            ("AZR-280", "エージェントを外部から停止する独立サーキットブレーカーを実装している", "モデル判断に依存せず、管理者・監視ルールから即時停止できる"),
            ("AZR-281", "高影響アクション前にhuman-in-the-loopチェックポイントを強制している", "対象操作、承認者、期限、拒否・無応答時動作がワークフローで強制される"),
            ("AZR-282", "エージェント異常を人間へエスカレーションする条件と経路を実装している", "反復、低信頼度、ポリシー違反、依存障害の通知と引継ぎがテストされている"),
        ],
        "AZR-249": [
            ("AZR-283", "ユーザー入力をコンテンツ安全性サービスで検査している", "全入力経路でカテゴリ別の検知・拒否・記録が動作する"),
            ("AZR-284", "モデル出力をコンテンツ安全性サービスで検査している", "全出力経路で違反内容が利用者・ツールへ渡らない"),
            ("AZR-285", "画像等のマルチモーダル入力に埋め込まれた有害・隠れ命令を検査している", "採用モダリティごとの攻撃・有害内容テストがある"),
            ("AZR-286", "複数モデル呼出しの安全性検査をゲートウェイ／オーケストレーション層へ集約している", "モデル直結の迂回経路がなく、共通ポリシーの性能・費用影響を測定している"),
        ],
    }
    for parent, pieces in azure_splits.items():
        parent_row = find_row(azure, parent)
        original_ref = azure.cell(parent_row, 8).value
        set_item(
            azure,
            parent,
            check=f"{parent}に関する統制範囲・責任者・適用条件が一つの設計判断として定義されている",
            criterion="対象コンポーネント、責任者、適用・非適用条件、関連する個別統制IDが文書化されている",
            intent="複合項目を個別に判定できるサブ統制へ分解し、親項目は適用範囲の統制に限定する。",
            reference=original_ref,
        )
        insert_items(
            azure,
            parent,
            [
                {
                    "id": item_id,
                    "check": check,
                    "criterion": criterion,
                    "intent": f"{check.rstrip('。')}ことを独立に判定し、部分実装を一括Passにしない。",
                    "reference": original_ref,
                }
                for item_id, check, criterion in pieces
            ],
        )


SOURCE_ROWS = [
    ("SRC-SWEBOK-V4A", "Guide to the Software Engineering Body of Knowledge (SWEBOK Guide) v4.0a", "v4.0a / 2025-09（ローカル固定参照版）", "https://www.computer.org/education/bodies-of-knowledge/software-engineering", "全18KAを対応付け、KA 1〜10・12・13を直接参照、KA 11・14・15を部分整合、KA 16〜18を理由付き対象外とする。"),
    ("SRC-AWS-WAF", "AWS Well-Architected Framework", "継続更新", "https://docs.aws.amazon.com/wellarchitected/latest/framework/welcome.html", "AWSおよびクラウド共通"),
    ("SRC-AWS-GENAI", "AWS Well-Architected Generative AI Lens", "継続更新", "https://docs.aws.amazon.com/wellarchitected/latest/generative-ai-lens/generative-ai-lens.html", "生成AI"),
    ("SRC-AWS-RAI", "AWS Well-Architected Responsible AI Lens", "2025-11-19", "https://docs.aws.amazon.com/wellarchitected/latest/responsible-ai-lens/responsible-ai-lens.html", "責任あるAI"),
    ("SRC-AWS-ML", "AWS Well-Architected Machine Learning Lens", "2025-11-19", "https://docs.aws.amazon.com/wellarchitected/latest/machine-learning-lens/machine-learning-lens.html", "従来型MLを含むML全般"),
    ("SRC-AWS-AGENTIC", "AWS Well-Architected Agentic AI Lens", "2026-06-10", "https://docs.aws.amazon.com/wellarchitected/latest/agentic-ai-lens/agentic-ai-lens.html", "エージェントAI"),
    ("SRC-GCP-WAF", "Google Cloud Well-Architected Framework", "継続更新", "https://cloud.google.com/architecture/framework", "Google Cloudおよびクラウド共通"),
    ("SRC-GCP-AIML", "Google Cloud Well-Architected Framework: AI and ML perspective", "2024-10-11 review", "https://cloud.google.com/architecture/framework/perspectives/ai-ml", "AI/ML"),
    ("SRC-AZR-WAF", "Microsoft Azure Well-Architected Framework", "継続更新", "https://learn.microsoft.com/en-us/azure/well-architected/", "Azureおよびクラウド共通"),
    ("SRC-AZR-AI", "AI Workload Documentation - Azure Well-Architected Framework", "継続更新", "https://learn.microsoft.com/en-us/azure/well-architected/ai/", "Azure AI/ML"),
    ("SRC-OCI-WAF", "Well-architected framework for Oracle Cloud Infrastructure", "F29550-09 / 2025-05", "https://docs.oracle.com/en/solutions/oci-best-practices/toc.htm", "OCIおよびクラウド共通"),
    ("SRC-ISO-29148", "ISO/IEC/IEEE 29148 Systems and software engineering — Life cycle processes — Requirements engineering", "参照記載版", "https://www.iso.org/standard/72089.html", "要求工学の実務補完"),
    ("SRC-ISO-25010", "ISO/IEC 25010 Systems and software Quality Requirements and Evaluation", "参照記載版", "https://www.iso.org/standard/78176.html", "品質モデルの実務補完"),
    ("SRC-ISO-GENERAL", "ISO/IEC/IEEE standards（項目の参照箇所に規格番号を記載）", "各参照記載版", "https://www.iso.org/standards.html", "その他ISO/IEC/IEEE規格"),
    ("SRC-OWASP", "OWASP application security guidance", "継続更新", "https://owasp.org/www-project-application-security-verification-standard/", "ASVS、Top 10等"),
    ("SRC-NIST", "NIST cybersecurity and AI guidance", "各参照記載版", "https://www.nist.gov/itl", "SSDF、AI RMF等"),
    ("SRC-CIS", "CIS Benchmarks", "継続更新", "https://www.cisecurity.org/cis-benchmarks", "構成・セキュリティベースライン"),
    ("SRC-RFC", "IETF RFCs", "各参照記載版", "https://www.rfc-editor.org/", "RFC 2119等"),
    ("SRC-SRE", "Google Site Reliability Engineering resources", "継続更新", "https://sre.google/books/", "SRE実務補完"),
    ("SRC-DORA", "DORA research and guidance", "継続更新", "https://dora.dev/", "デリバリ性能の実務補完"),
    ("SRC-FINOPS", "FinOps Framework", "継続更新", "https://www.finops.org/framework/", "クラウドコスト管理"),
    ("SRC-IEEE-GENERAL", "IEEE standards and guidance（項目の参照箇所に規格番号を記載）", "各参照記載版", "https://standards.ieee.org/", "その他IEEE参照"),
    ("SRC-PRACTICE", "組織・案件固有の実務補完", "案件ごとに根拠を登録", "", "外部規範ではない。証跡欄に組織標準・意思決定記録を添付する。"),
]

SOURCE_CHANGE_SUMMARY = {
    "SRC-SWEBOK-V4A": "v4.0から2025年9月の軽微改訂版v4.0aへ更新。完全準拠ではなく対象範囲を明示した参照・整合へ位置付けを修正。",
    "SRC-AWS-GENAI": "2025年11月19日公開の現行Generative AI Lensを生成AI用途へ対応付け。",
    "SRC-AWS-RAI": "2025年11月19日公開のResponsible AI Lensを追加し、案件固有の考慮事項として扱う。",
    "SRC-AWS-ML": "2025年11月19日公開のMachine Learning Lensを従来型ML用途へ追加。",
    "SRC-AWS-AGENTIC": "2026年6月10日公開のAgentic AI Lensを追加し、自律性・追跡・監督の統制へ対応付け。",
    "SRC-GCP-AIML": "AI/ML perspectiveを一般WAFから分離してAI条件付きプロファイルへ対応付け。",
    "SRC-AZR-AI": "Azure AI Workload資料を一般WAFから分離してAI条件付きプロファイルへ対応付け。",
}


def source_ids(sheet: str, reference: str) -> str:
    text = reference or ""
    low = text.lower()
    ids: list[str] = []

    def add(value: str) -> None:
        if value not in ids:
            ids.append(value)

    if "swebok-v4.pdf" in low or "swebok-v4a.pdf" in low:
        add("SRC-SWEBOK-V4A")
    if "responsible ai lens" in low:
        add("SRC-AWS-RAI")
    if "agentic ai lens" in low:
        add("SRC-AWS-AGENTIC")
    if "machine learning lens" in low:
        add("SRC-AWS-ML")
    if "generative ai lens" in low or "genai lens" in low or "gensec" in low or "genops" in low or "genrel" in low or "gencost" in low or "genperf" in low:
        add("SRC-AWS-GENAI")
    if "aws" in low and not any(value.startswith("SRC-AWS-") for value in ids):
        add("SRC-AWS-WAF")
    if "google" in low or "gcp" in low:
        add("SRC-GCP-AIML" if ("ai" in low or "ml" in low or sheet == "19_AI") else "SRC-GCP-WAF")
    if "azure" in low:
        add("SRC-AZR-AI" if (sheet == "19_AI" or (sheet == "17_Azure" and any(token in low for token in ["ai workload", "model", "grounding", "agent", "training data"]))) else "SRC-AZR-WAF")
    if "oci" in low or "oracle cloud" in low:
        add("SRC-OCI-WAF")
    if "29148" in low:
        add("SRC-ISO-29148")
    if "25010" in low:
        add("SRC-ISO-25010")
    if "iso" in low and not any(value.startswith("SRC-ISO-") for value in ids):
        add("SRC-ISO-GENERAL")
    if "owasp" in low:
        add("SRC-OWASP")
    if "nist" in low:
        add("SRC-NIST")
    if "cis" in low:
        add("SRC-CIS")
    if "rfc" in low:
        add("SRC-RFC")
    if "sre" in low:
        add("SRC-SRE")
    if "dora" in low:
        add("SRC-DORA")
    if "finops" in low:
        add("SRC-FINOPS")
    if "ieee" in low and "SRC-SWEBOK-V4A" not in ids and not any(value.startswith("SRC-ISO-") for value in ids):
        add("SRC-IEEE-GENERAL")
    if "実務補完" in text or not ids:
        add("SRC-PRACTICE")
    return ";".join(ids)


def duplicate_group(check: str, category: str) -> str:
    text = f"{check} {category}".lower()
    rules = [
        (["バックアップ", "backup", "復元"], "CONTINUITY-BACKUP"),
        (["rto", "rpo", "dr", "災害復旧", "ディザスタ", "切替"], "CONTINUITY-DR"),
        (["可用性", "冗長", "availability zone", "障害ドメイン", "フェイルオーバー"], "RELIABILITY-HA"),
        (["最小権限", "iam", "rbac", "abac", "認証", "認可", "特権"], "SEC-IAM"),
        (["暗号", "鍵管理", "key vault", "kms"], "SEC-CRYPTO"),
        (["監視", "ログ", "メトリクス", "トレース", "アラート", "observability"], "OPS-OBSERVABILITY"),
        (["iac", "infrastructure as code"], "OPS-IAC"),
        (["sbom", "サプライチェーン", "依存ライブラリ", "脆弱性"], "SEC-SUPPLY-CHAIN"),
        (["個人データ", "個人情報", "プライバシ", "pii", "忘れられる権利"], "PRIVACY"),
        (["プロンプトインジェクション", "ジェイルブレイク"], "AI-PROMPT-SEC"),
        (["コスト", "費用", "予算", "finops"], "COST"),
    ]
    return ";".join(group for words, group in rules if any(word in text for word in words))


def applicability(sheet: str, item_id: str, check: str) -> str:
    profile = PROFILE_BY_SHEET[sheet]
    if item_id == "CLD-001":
        return "可用性SLO、許容停止時間、障害ドメイン分析が冗長化を必要とする場合。非適用・代替は根拠を記録。"
    if item_id in {"CLD-004", "CLD-065", "CLD-066"}:
        return "事業影響分析とRTO/RPOでDRが必要な場合。クロスリージョンはリージョン障害対策が要求される場合のみ。"
    if item_id == "CLD-011":
        return "水平スケール、ローリング更新、フェイルオーバー対象のコンポーネント。"
    if item_id == "AI-007":
        return "個人データを処理する場合。適法根拠・同意要否は適用法令と組織ポリシーによる。"
    if item_id == "AI-041":
        return "AI機能のSLO・RTO/RPOに対して単一障害点対策が必要な場合。冗長化方式は要求とコストで選定。"
    defaults = {
        "CORE": "原則適用。案件スコープ・成果物・規制・リスク上非該当の場合はN/A根拠を記録。",
        "CLOUD-COMMON": "クラウドを利用する案件に適用。サービス特性・責任共有モデルに応じてN/A可。",
        "AWS-DELTA": "AWS採用時に、CLOUD-COMMONで未評価のAWS固有差分へ適用。",
        "GCP-DELTA": "Google Cloud採用時に、CLOUD-COMMONで未評価の固有差分へ適用。",
        "AZURE-DELTA": "Azure採用時に、CLOUD-COMMONで未評価のAzure固有差分へ適用。",
        "OCI-DELTA": "OCI採用時に、CLOUD-COMMONで未評価のOCI固有差分へ適用。",
        "AI-CONDITIONAL": "AI/MLを含む案件に適用。従来型ML・生成AI・エージェントAIの種別ごとにN/Aを判断。",
    }
    return defaults[profile]


def extend_review_sheets(wb) -> None:
    for sheet in REVIEW_SHEETS:
        ws = wb[sheet]
        # Idempotent rebuild of the execution/source extension.
        if ws.max_column > 8:
            ws.delete_cols(9, ws.max_column - 8)
        ws.cell(1, 4).value = "基準重要度"
        header_style = copy(ws.cell(1, 8)._style)
        for offset, header in enumerate(EXT_HEADERS, 9):
            cell = ws.cell(1, offset, header)
            cell._style = copy(header_style)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in range(2, ws.max_row + 1):
            item_id = str(ws.cell(row, 1).value or "")
            check = str(ws.cell(row, 5).value or "")
            category = str(ws.cell(row, 3).value or "")
            reference = str(ws.cell(row, 8).value or "")
            base_style = copy(ws.cell(row, 8)._style)
            for col in range(9, 24):
                ws.cell(row, col)._style = copy(base_style)
                ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row, 9).value = PROFILE_BY_SHEET[sheet]
            ws.cell(row, 10).value = applicability(sheet, item_id, check)
            ws.cell(row, 11).value = source_ids(sheet, reference)
            ws.cell(row, 12).value = duplicate_group(check, category)
        ws.auto_filter.ref = f"A1:W{ws.max_row}"
        ws.freeze_panes = "E2"
        widths = {
            "A": 14, "B": 14, "C": 28, "D": 12, "E": 46, "F": 52, "G": 48, "H": 46,
            "I": 18, "J": 38, "K": 28, "L": 25, "M": 13, "N": 12, "O": 12, "P": 38,
            "Q": 24, "R": 30, "S": 23, "T": 15, "U": 18, "V": 15, "W": 15,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        max_rows = max(ws.max_row + 500, 2000)
        validations = [
            (DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True), f"M2:M{max_rows}"),
            (DataValidation(type="list", formula1='"未判定,適用,N/A"', allow_blank=True), f"N2:N{max_rows}"),
            (DataValidation(type="list", formula1='"未確認,Pass,Fail"', allow_blank=True), f"O2:O{max_rows}"),
            (DataValidation(type="list", formula1='"未確認,Pass,Fail"', allow_blank=True), f"W2:W{max_rows}"),
        ]
        for dv, area in validations:
            ws.add_data_validation(dv)
            dv.add(area)
        red = PatternFill("solid", fgColor="FFC7CE")
        green = PatternFill("solid", fgColor="C6EFCE")
        gray = PatternFill("solid", fgColor="D9E1F2")
        ws.conditional_formatting.add(f"O2:O{max_rows}", CellIsRule(operator="equal", formula=['"Fail"'], fill=red))
        ws.conditional_formatting.add(f"O2:O{max_rows}", CellIsRule(operator="equal", formula=['"Pass"'], fill=green))
        ws.conditional_formatting.add(f"N2:N{max_rows}", CellIsRule(operator="equal", formula=['"N/A"'], fill=gray))
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.print_title_rows = "1:1"


NAVY = "17365D"
BLUE = "4472C4"
LIGHT_BLUE = "D9EAF7"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="B7C9E2")


def title_cell(ws, title: str, end_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    cell = ws.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(color=WHITE, bold=True, size=16)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28


def style_table(ws, header_row: int, start_col: int, end_col: int, end_row: int) -> None:
    for cell in ws[header_row][start_col - 1:end_col]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    for row in ws.iter_rows(min_row=header_row + 1, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_guide(wb, swebok_hash: str) -> None:
    if "利用ガイド" in wb.sheetnames:
        del wb["利用ガイド"]
    ws = wb.create_sheet("利用ガイド", 1)
    title_cell(ws, "利用ガイド — 適用、判定、重複統制、リリースゲート", 7)
    rows = [
        (3, "位置づけ", "SWEBOK Guide v4.0aと各社Well-Architected Frameworkを参照・整合させたレビュー観点マスター兼実施記録。規格認証や全SWEBOK KAへの完全準拠を表明するものではない。"),
        (4, "SWEBOK参照版", f"ローカルの .workspace/swebok-v4a.pdf を固定参照版とする（SHA-256: {swebok_hash}）。PDF本体は再配布しない。"),
        (5, "SWEBOK対象範囲", "全18 KAを対応付ける。KA 1〜10・12・13は直接参照、KA 11・14・15は部分整合、KA 16〜18は前提知識として理由付き対象外。詳細は「KA対応表」を参照。"),
        (6, "WAFの扱い", "WAFは設計判断・リスク・トレードオフを評価する助言資料。各項目は案件のSLO、RTO/RPO、規制、データ分類、費用制約を入力に適用判定する。"),
    ]
    for row, label, value in rows:
        ws.cell(row, 1, label).font = Font(bold=True, color=NAVY)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        ws.cell(row, 2, value).alignment = Alignment(wrap_text=True, vertical="top")

    ws.cell(8, 1, "適用プロファイル")
    ws.cell(9, 1, "プロファイル")
    ws.cell(9, 2, "適用順")
    ws.cell(9, 3, "対象")
    ws.cell(9, 4, "利用シート")
    ws.cell(9, 5, "重複時の扱い")
    profiles = [
        ("CORE", 1, "原則全案件", "01〜10、12、13", "まず評価する"),
        ("CLOUD-COMMON", 2, "クラウド利用案件", "14_クラウド", "Coreの証跡を再利用可"),
        ("AWS/GCP/AZURE/OCI-DELTA", 3, "採用ベンダーのみ", "15〜18の該当1シート", "共通で未評価の固有差分を追加"),
        ("AI-CONDITIONAL", 4, "AI/ML利用案件", "19_AI", "従来ML・生成AI・Agenticの種別でN/A判断"),
    ]
    for r, values in enumerate(profiles, 10):
        for c, value in enumerate(values, 1):
            ws.cell(r, c, value)
    style_table(ws, 9, 1, 5, 13)

    ws.cell(15, 1, "レビュー実施手順")
    workflow = [
        "1. 適用プロファイルを選び、各行の「適用判定」を適用／N/Aで記録する。N/Aは「証跡・参照箇所」に根拠を残す。",
        "2. 基準重要度を助言値として、案件の重要度・データ分類・規制・SLOを反映した「案件重要度」と判断根拠を設定する。",
        "3. Pass／Fail／未確認を記録し、Passは証跡、Failは指摘・Issue ID、対応方針、是正期限、担当情報を残す。",
        "4. N/Aにもレビュアーと日付を残す。Fail是正後は明示的な再確認記録によりPassを確認し、変更履歴を保持する。",
    ]
    for r, value in enumerate(workflow, 16):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws.cell(r, 1, value)
        ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")

    ws.cell(21, 1, "重要度とゲート")
    ws.cell(22, 1, "区分")
    ws.cell(22, 2, "定義")
    ws.cell(22, 3, "運用")
    severity_rows = [
        ("基準重要度", "観点が一般に持つ潜在影響。マスター側の初期値。", "案件固有のリリース判定を単独では決めない。"),
        ("案件重要度", "対象システムの重要度、規制、データ、SLO、露出度を反映した実行時評価。", "レビュー責任者が根拠とともに設定する。"),
        ("案件Critical", "適用かつ未解決のFailが重大事故・法令違反・受入不能へ直結する。", "原則リリース停止。例外は記名承認と期限付きリスク受容が必要。"),
    ]
    for r, values in enumerate(severity_rows, 23):
        for c, value in enumerate(values, 1):
            ws.cell(r, c, value)
    style_table(ws, 22, 1, 3, 25)

    ws.cell(27, 1, "重複統制ルール")
    duplicate_rules = [
        "同じリスクがCore、Cloud Common、ベンダー、AIに現れる場合、先に実施した証跡を再利用してよい。",
        "共通統制で十分に評価済みのベンダー行はN/Aとし、「証跡・参照箇所」に評価済みIDを記載する。ベンダー固有差分がある場合だけ別判定する。",
        "同じ根本リスクは同一のIssue IDを使用し、件数・リスク集計ではIssue IDを一意に数える。行数の合算をリスク件数にしない。",
        "「重複統制グループ」は候補を機械付与した補助情報。最終的な同一統制判断は証跡・対象資産・要求を確認して行う。",
    ]
    for r, value in enumerate(duplicate_rules, 28):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws.cell(r, 1, f"• {value}")
        ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")

    ws.cell(33, 1, "追加承認の移行注意")
    ws.merge_cells(start_row=34, start_column=1, end_row=35, end_column=7)
    ws.cell(34, 1, "旧ブックで「追加承認」とされた65件は「追加項目管理」へ全件接続した。ただし旧ブックには承認者・承認日の原記録がないため、状態を「要原記録確認」としている。利用開始前に実在する承認記録で補完し、未確認のまま承認済みと扱わない。")
    ws.cell(34, 1).fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(34, 1).alignment = Alignment(wrap_text=True, vertical="top")
    widths = {"A": 28, "B": 48, "C": 48, "D": 26, "E": 38, "F": 22, "G": 22}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A3"


def build_source_master(wb, swebok_hash: str) -> None:
    if "出典マスター" in wb.sheetnames:
        del wb["出典マスター"]
    ws = wb.create_sheet("出典マスター", 2)
    title_cell(ws, "出典マスター", 10)
    headers = ["Source ID", "正式名称", "版・公開日", "URL", "参照日", "適用範囲", "変更確認日", "前版との差分", "固定参照物", "注記"]
    for c, value in enumerate(headers, 1):
        ws.cell(3, c, value)
    for r, (source_id, name, version, url, scope) in enumerate(SOURCE_ROWS, 4):
        fixed = f".workspace/swebok-v4a.pdf / SHA-256 {swebok_hash}" if source_id == "SRC-SWEBOK-V4A" else ""
        note = "添付物の同一性だけを固定し、PDF本体は再配布しない。" if source_id == "SRC-SWEBOK-V4A" else "継続更新型資料は変更確認日に再確認する。"
        change_summary = SOURCE_CHANGE_SUMMARY.get(source_id, "前回登録内容から適用範囲・版の変更なし。")
        values = [source_id, name, version, url, UPDATED, scope, UPDATED, change_summary, fixed, note]
        for c, value in enumerate(values, 1):
            ws.cell(r, c, value)
        if url:
            ws.cell(r, 4).hyperlink = url
            ws.cell(r, 4).style = "Hyperlink"
    style_table(ws, 3, 1, 10, 3 + len(SOURCE_ROWS))
    for col, width in {"A": 23, "B": 48, "C": 24, "D": 65, "E": 14, "F": 48, "G": 14, "H": 58, "I": 68, "J": 42}.items():
        ws.column_dimensions[col].width = width
    ws.auto_filter.ref = f"A3:J{ws.max_row}"
    ws.freeze_panes = "A4"


def build_ka_map(wb) -> None:
    if "KA対応表" in wb.sheetnames:
        del wb["KA対応表"]
    ws = wb.create_sheet("KA対応表", 3)
    title_cell(ws, "SWEBOK v4.0a Knowledge Area 対応表", 7)
    headers = ["KA", "Knowledge Area", "扱い", "対応シート", "適用範囲・除外根拠", "主な証跡", "備考"]
    for c, value in enumerate(headers, 1):
        ws.cell(3, c, value)
    kas = [
        (1, "Software Requirements", "直接参照", "01_要件定義", "レビュー対象", "要求仕様・トレーサビリティ", ""),
        (2, "Software Architecture", "直接参照", "02_アーキテクチャ", "レビュー対象", "アーキテクチャ記述・意思決定", ""),
        (3, "Software Design", "直接参照", "03_詳細設計", "レビュー対象", "詳細設計・API/DB/UI設計", ""),
        (4, "Software Construction", "直接参照", "04_実装", "レビュー対象", "コード・実装規約・ビルド", ""),
        (5, "Software Testing", "直接参照", "05_テスト", "レビュー対象", "テスト計画・設計・結果", ""),
        (6, "Software Engineering Operations", "直接参照", "06_運用", "レビュー対象", "運用設計・Runbook・SLO", ""),
        (7, "Software Maintenance", "直接参照", "07_保守", "レビュー対象", "保守計画・変更・廃止", ""),
        (8, "Software Configuration Management", "直接参照", "08_構成管理", "レビュー対象", "構成識別・変更・リリース", ""),
        (9, "Software Engineering Management", "直接参照", "09_マネジメント", "レビュー対象", "計画・見積・リスク・測定", ""),
        (10, "Software Engineering Process", "直接参照", "10_プロセス", "レビュー対象", "プロセス定義・評価・改善", ""),
        (11, "Software Engineering Models and Methods", "部分整合（独立シートなし）", "02/03/05/10/12", "モデル・形式手法・レビュー方法を各工程で条件付き評価。方法論全体の網羅は対象外。", "モデル、分析・レビュー記録", "独立KA準拠は表明しない"),
        (12, "Software Quality", "直接参照", "12_品質保証", "レビュー対象", "品質計画・保証・測定", ""),
        (13, "Software Security", "直接参照", "13_セキュリティ", "レビュー対象", "脅威モデル・セキュリティ設計/試験", ""),
        (14, "Software Engineering Professional Practice", "部分整合（独立シートなし）", "01/09/10/12/19", "コミュニケーション、倫理、責任あるAIの成果物化された部分のみ。専門職実践全体は対象外。", "合意記録・役割・倫理審査", "独立KA準拠は表明しない"),
        (15, "Software Engineering Economics", "部分整合（独立シートなし）", "09/14〜19", "見積、投資、クラウド/AIコスト、トレードオフのみ。経済学体系全体は対象外。", "Business case・TCO・予算", "独立KA準拠は表明しない"),
        (16, "Computing Foundations", "対象外（前提知識）", "—", "個別成果物レビューではなく実務者の基礎知識。必要な技術事項は各工程で評価。", "—", "適用除外"),
        (17, "Mathematical Foundations", "対象外（前提知識）", "—", "個別成果物レビューではなく基礎知識。安全・形式検証案件では条件付きで方法を評価。", "—", "適用除外"),
        (18, "Engineering Foundations", "対象外（前提知識）", "—", "個別成果物レビューではなく工学基礎。安全・規制領域は案件プロファイルで補完。", "—", "適用除外"),
    ]
    for r, values in enumerate(kas, 4):
        for c, value in enumerate(values, 1):
            ws.cell(r, c, value)
    style_table(ws, 3, 1, 7, 3 + len(kas))
    for col, width in {"A": 8, "B": 42, "C": 27, "D": 24, "E": 58, "F": 38, "G": 28}.items():
        ws.column_dimensions[col].width = width
    ws.auto_filter.ref = f"A3:G{ws.max_row}"
    ws.freeze_panes = "A4"


def build_addition_register(wb) -> None:
    if "追加項目管理" in wb.sheetnames:
        del wb["追加項目管理"]
    ws = wb.create_sheet("追加項目管理", 4)
    title_cell(ws, "追加承認項目の由来・承認記録", 11)
    headers = ["追加ID", "シート", "追加理由／不足トピック", "追加チェック項目", "根拠資料", "出典ID", "承認状態", "承認者", "承認日", "変更理由", "注記"]
    for c, value in enumerate(headers, 1):
        ws.cell(3, c, value)
    gap = wb["抜け漏れ分析"]
    gap_map = {}
    for row in gap.iter_rows(min_row=2, values_only=True):
        if row[3]:
            gap_map[str(row[3])] = row
    records = []
    for sheet in REVIEW_SHEETS:
        source_ws = wb[sheet]
        for r in range(2, source_ws.max_row + 1):
            if source_ws.cell(r, 2).value != "追加承認":
                continue
            item_id = str(source_ws.cell(r, 1).value)
            gap_row = gap_map.get(item_id)
            reason = str(gap_row[1]) if gap_row else f"旧ブックで追加承認として登録（カテゴリ: {source_ws.cell(r, 3).value}）"
            change_reason = str(gap_row[2]) if gap_row else "旧ブックに追加理由・原承認記録なし。原記録の確認が必要。"
            records.append([
                item_id,
                sheet,
                reason,
                source_ws.cell(r, 5).value,
                source_ws.cell(r, 8).value,
                source_ws.cell(r, 11).value,
                "要原記録確認",
                "未記録",
                "未記録",
                change_reason,
                "旧ブックの「追加承認」表記を移行。承認者・日付を推測で補完しない。",
            ])
    records.sort(key=lambda row: (REVIEW_SHEETS.index(row[1]), row[0]))
    for r, values in enumerate(records, 4):
        for c, value in enumerate(values, 1):
            ws.cell(r, c, value)
    style_table(ws, 3, 1, 11, 3 + len(records))
    for col, width in {"A": 16, "B": 22, "C": 48, "D": 48, "E": 48, "F": 26, "G": 18, "H": 18, "I": 15, "J": 55, "K": 48}.items():
        ws.column_dimensions[col].width = width
    ws.auto_filter.ref = f"A3:K{ws.max_row}"
    ws.freeze_panes = "A4"


def update_summary(wb, swebok_hash: str) -> None:
    ws = wb["サマリ"]
    ws["A1"] = "SWEBOK Guide v4.0a（固定参照版）・クラウドWell-Architected Framework参照・整合 開発レビューチェックリスト"
    ws["A2"] = (
        f"更新: {UPDATED.isoformat()} / SWEBOK参照版: v4.0a / .workspace/swebok-v4a.pdf (SHA-256: {swebok_hash})。"
        "全18KAとの関係と適用・除外理由は「KA対応表」を参照。WAFは規範的な一律要件ではなく、案件条件に基づき適用判定する。"
        "各レビューシートに適用／N/A、Pass／Fail、証跡、指摘、例外承認、再確認の実施記録列を追加。"
        "出典は「出典マスター」、運用は「利用ガイド」、追加承認65件は「追加項目管理」を参照。"
    )
    ws.cell(4, 4).value = "項目数（数式）"
    for row, sheet in enumerate(REVIEW_SHEETS, 5):
        q = quote_sheetname(sheet)
        end_row = wb[sheet].max_row
        ws.cell(row, 4).value = f"=COUNTA({q}!$A$2:$A${end_row})"
        for col, severity in zip(range(5, 9), ["Critical", "High", "Medium", "Low"]):
            ws.cell(row, col).value = f'=COUNTIF({q}!$D$2:$D${end_row},"{severity}")'
    total_row = 5 + len(REVIEW_SHEETS)
    ws.cell(total_row, 1).value = "合計"
    for col in range(4, 9):
        letter = get_column_letter(col)
        ws.cell(total_row, col).value = f"=SUM({letter}5:{letter}{total_row - 1})"
    ws["I23"] = "件数は基準重要度の集計。案件重要度・判定結果は各レビュー実施時に記録する。"
    ws.column_dimensions["I"].width = 70
    ws.freeze_panes = "A5"


def verify(wb, swebok_hash: str) -> None:
    assert swebok_hash == EXPECTED_SWEBOK_SHA256, "unexpected SWEBOK v4.0a artifact"
    all_ids: list[str] = []
    source_set = {row[0] for row in SOURCE_ROWS}
    for sheet in REVIEW_SHEETS:
        ws = wb[sheet]
        assert ws.max_column == 23, (sheet, ws.max_column)
        assert [ws.cell(1, c).value for c in range(9, 24)] == EXT_HEADERS
        for row in range(2, ws.max_row + 1):
            item_id = ws.cell(row, 1).value
            assert item_id, (sheet, row, "missing ID")
            all_ids.append(str(item_id))
            for col in range(1, 12):
                assert ws.cell(row, col).value not in (None, ""), (sheet, row, col)
            for source in str(ws.cell(row, 11).value).split(";"):
                assert source in source_set, (sheet, item_id, source)
    assert len(all_ids) == len(set(all_ids)), "duplicate IDs"
    addition_ws = wb["追加項目管理"]
    assert addition_ws.max_row - 3 == 65, addition_ws.max_row - 3
    assert len(wb["KA対応表"]["A"]) - 3 == 18
    assert wb["出典マスター"].max_column == 10
    assert wb["出典マスター"].cell(3, 8).value == "前版との差分"
    for row in wb["サマリ"].iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                assert "1048576" not in cell.value, cell.coordinate
            if isinstance(cell.value, str):
                assert "swebok-v4.pdf" not in cell.value, cell.coordinate
    for required in ["CLD-065", "CLD-066", "AI-080", "AI-076", "AZR-286"]:
        assert required in all_ids, required


def main() -> None:
    if not BOOK.exists() or not SWEBOK.exists():
        raise SystemExit(f"checklist.xlsx and {SWEBOK} are required")
    swebok_hash = sha256(SWEBOK)
    if swebok_hash != EXPECTED_SWEBOK_SHA256:
        raise SystemExit(f"unexpected SWEBOK v4.0a SHA-256: {swebok_hash}")
    wb = load_workbook(BOOK)
    for generated in ["利用ガイド", "出典マスター", "KA対応表", "追加項目管理"]:
        if generated in wb.sheetnames:
            del wb[generated]
    normalize_swebok_references(wb)
    apply_content_fixes(wb)
    extend_review_sheets(wb)
    build_guide(wb, swebok_hash)
    build_source_master(wb, swebok_hash)
    build_ka_map(wb)
    build_addition_register(wb)
    update_summary(wb, swebok_hash)
    verify(wb, swebok_hash)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.properties.title = "SWEBOK Guide v4.0a・クラウドWell-Architected参照 開発レビューチェックリスト"
    wb.properties.subject = "レビュー観点マスター・レビュー実施記録"
    wb.properties.modified = datetime.combine(UPDATED, datetime.min.time())
    stale_temp = BOOK.with_name(f".{BOOK.name}.tmp")
    if stale_temp.exists():
        stale_temp.unlink()
    temp = BOOK.with_name(f".{BOOK.stem}.tmp{BOOK.suffix}")
    wb.save(temp)
    # Reopen the serialized artifact before replacing the original.
    check = load_workbook(temp, data_only=False)
    verify(check, swebok_hash)
    check.close()
    os.replace(temp, BOOK)
    print(f"updated {BOOK} / SWEBOK SHA-256 {swebok_hash}")


if __name__ == "__main__":
    main()
