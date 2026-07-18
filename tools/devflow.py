#!/usr/bin/env python3
"""AI開発のライフサイクルを決定的に統制する実行基盤。"""

from __future__ import annotations

import argparse
import hashlib
import importlib.util
import json
import os
import re
import subprocess
import sys
import tempfile
import uuid
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

ROOT = Path(__file__).resolve().parents[1]
POLICY_PATH = ROOT / "governance" / "policy.json"
CATALOG_PATH = ROOT / "governance" / "checklist" / "catalog.json"
WORK_ROOT = ROOT / "work"
TEMPLATE_ROOT = ROOT / "docs" / "templates"
IMPROVEMENTS_PATH = ROOT / "governance" / "improvements" / "proposals.json"
REPORT_ROOT = ROOT / "reports"
RESULTS_SCHEMA_VERSION = 2
SCOPEFLOW_PATH = ROOT / ".agents" / "skills" / "right-size-execution" / "scripts" / "scopeflow.py"


class GovernanceError(RuntimeError):
    pass


def utcnow() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def read_json(path: Path) -> Any:
    try:
        display = path.relative_to(ROOT)
    except ValueError:
        display = path
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise GovernanceError(f"required file not found: {display}") from exc
    except json.JSONDecodeError as exc:
        raise GovernanceError(f"invalid JSON: {display}: {exc}") from exc


def canonical_bytes(value: Any) -> bytes:
    return (json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")) + "\n").encode("utf-8")


def atomic_write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(value, ensure_ascii=False, sort_keys=True, indent=2) + "\n"
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", dir=path.parent, delete=False) as stream:
        stream.write(payload)
        stream.flush()
        os.fsync(stream.fileno())
        temp = Path(stream.name)
    os.replace(temp, path)
    path.chmod(0o644)


def sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def load_policy() -> dict[str, Any]:
    return read_json(POLICY_PATH)


def load_catalog() -> dict[str, Any]:
    return read_json(CATALOG_PATH)


def safe_work_path(work_item: str) -> Path:
    if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._-]{2,120}", work_item):
        raise GovernanceError("work item ID contains unsupported characters")
    path = (WORK_ROOT / work_item).resolve()
    if WORK_ROOT.resolve() not in path.parents:
        raise GovernanceError("work item path escapes work root")
    if not path.is_dir():
        raise GovernanceError(f"work item does not exist: {work_item}")
    return path


def infer_phase(sheet: str, category: str, check: str) -> str:
    direct = {
        "01_要件定義": "requirements",
        "02_アーキテクチャ": "architecture",
        "03_詳細設計": "design",
        "04_実装": "implementation",
        "05_テスト": "verification",
        "06_運用": "operations",
        "07_保守": "operations",
        "08_構成管理": "implementation",
        "09_マネジメント": "release",
        "10_プロセス": "release",
        "12_品質保証": "verification",
    }
    if sheet in direct:
        return direct[sheet]
    text = f"{category} {check}".lower()
    if sheet == "13_セキュリティ":
        if any(word in text for word in ["テスト", "検証", "脆弱性", "ペネトレーション", "監査"]):
            return "verification"
        if any(word in text for word in ["実装", "コーディング", "ライブラリ", "依存"]):
            return "implementation"
        return "architecture"
    if sheet == "19_AI":
        if any(word in text for word in ["ユースケース", "ガバナンス", "ステークホルダー", "適法"]):
            return "requirements"
        if any(word in text for word in ["評価", "テスト", "レッドチーム", "ジェイルブレイク"]):
            return "verification"
        if any(word in text for word in ["デプロイ", "genaiops", "バージョン管理", "ci/cd"]):
            return "implementation"
        if any(word in text for word in ["運用", "監視", "ドリフト", "アラート", "廃止"]):
            return "operations"
        if any(word in text for word in ["データ", "モデル選定", "カスタマイズ", "グラウンディング"]):
            return "design"
        return "architecture"
    if any(word in text for word in ["テスト", "評価", "検証", "演習"]):
        return "verification"
    if any(word in text for word in ["運用", "監視", "ログ", "アラート", "コスト最適化", "バックアップ"]):
        return "operations"
    if any(word in text for word in ["デプロイ", "構成管理", "iac", "パイプライン"]):
        return "implementation"
    return "architecture"


def checklist_scope_attributes(sheet: str, category: str, check: str, phase: str) -> dict[str, Any]:
    """Derive deterministic selector attributes without adding review-result columns."""
    text = f"{sheet} {category} {check}".lower()
    artifact_rules = [
        (["要件", "要求", "受入基準", "トレーサビリティ"], "requirements"),
        (["設計", "アーキテクチャ"], "design"),
        (["api", "インターフェース", "契約"], "public-api"),
        (["データベース", "db", "sql", "スキーマ", "マイグレーション"], "database"),
        (["実装", "コード", "コーディング"], "implementation"),
        (["テスト", "検証", "評価"], "test"),
        (["文書", "ドキュメント"], "documentation"),
        (["iac", "cloudformation", "terraform", "cdk"], "iac"),
        (["依存", "ライブラリ", "sbom", "lock"], "dependency"),
        (["ガバナンス", "チェックリスト", "プロセス", "品質ゲート"], "governance"),
        (["生成", "ジェネレータ"], "generator"),
        (["運用", "監視", "ログ", "アラート"], "operations"),
    ]
    risk_rules = [
        (["セキュリティ", "脆弱性", "攻撃", "脅威"], "security"),
        (["認証"], "authentication"),
        (["認可", "権限", "最小権限"], "authorization"),
        (["データ削除", "データ損失"], "data-loss"),
        (["スキーマ", "マイグレーション"], "database-schema"),
        (["公開api", "外部api", "イベント契約"], "public-api"),
        (["iac", "cloudformation", "terraform", "cdk", "ネットワーク", "デプロイ"], "iac"),
        (["依存", "ライブラリ", "sbom", "lock"], "dependency"),
        (["個人情報", "個人データ", "pii", "機密"], "pii"),
        (["ガバナンス", "チェックリスト", "正本", "生成器"], "governance"),
    ]
    artifacts = sorted({tag for words, tag in artifact_rules if any(word in text for word in words)})
    risks = sorted({tag for words, tag in risk_rules if any(word in text for word in words)})
    always_terms = ["受入基準", "トレーサビリティ", "完了条件", "変更管理", "証跡", "再現可能"]
    always_on = any(term in text for term in always_terms)
    if risks:
        scope_levels = [3]
    elif always_on or phase in {"implementation", "verification"}:
        scope_levels = [1, 2, 3]
    else:
        scope_levels = [2, 3]
    return {
        "always_on": always_on,
        "artifact_tags": artifacts,
        "risk_tags": risks,
        "scope_levels": scope_levels,
    }


def load_scopeflow() -> Any:
    if not SCOPEFLOW_PATH.is_file():
        raise GovernanceError(f"scopeflow not found: {SCOPEFLOW_PATH.relative_to(ROOT)}")
    spec = importlib.util.spec_from_file_location("dev_standard_scopeflow", SCOPEFLOW_PATH)
    if spec is None or spec.loader is None:
        raise GovernanceError("scopeflow module cannot be loaded")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def workbook_catalog() -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise GovernanceError("openpyxl is required; run: pip install -r requirements.txt") from exc

    policy = load_policy()
    book_path = ROOT / policy["standard"]["checklist"]
    if not book_path.exists():
        raise GovernanceError(f"checklist workbook not found: {book_path.relative_to(ROOT)}")
    sheets: list[str] = []
    for profile in policy["profiles"].values():
        for sheet in profile["sheets"]:
            if sheet not in sheets:
                sheets.append(sheet)
    wb = load_workbook(book_path, read_only=True, data_only=False)
    items: list[dict[str, Any]] = []
    for sheet in sheets:
        if sheet not in wb.sheetnames:
            raise GovernanceError(f"checklist sheet missing: {sheet}")
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        headers = [str(value or "") for value in next(rows)]
        index = {header: pos for pos, header in enumerate(headers)}
        required = ["ID", "区分", "カテゴリ", "基準重要度", "チェック項目", "合格基準", "意図(なぜ確認するか)", "対象ドキュメント(参照箇所)"]
        missing = [header for header in required if header not in index]
        if missing:
            raise GovernanceError(f"{sheet}: required columns missing: {', '.join(missing)}")
        for values in rows:
            item_id = values[index["ID"]]
            if not item_id:
                continue
            category = str(values[index["カテゴリ"]] or "")
            check = str(values[index["チェック項目"]] or "")
            phase = infer_phase(sheet, category, check)
            item = {
                "id": str(item_id),
                "sheet": sheet,
                "section": str(values[index["区分"]] or ""),
                "category": category,
                "base_severity": str(values[index["基準重要度"]] or ""),
                "check": check,
                "acceptance": str(values[index["合格基準"]] or ""),
                "intent": str(values[index["意図(なぜ確認するか)"]] or ""),
                "reference": str(values[index["対象ドキュメント(参照箇所)"]] or ""),
                "profile": str(values[index["適用プロファイル"]] or "") if "適用プロファイル" in index else "",
                "applicability_condition": str(values[index["適用条件"]] or "") if "適用条件" in index else "",
                "source_ids": [value for value in str(values[index["出典ID"]] or "").split(";") if value] if "出典ID" in index else [],
                "duplicate_groups": [value for value in str(values[index["重複統制グループ"]] or "").split(";") if value] if "重複統制グループ" in index else [],
                "phase": phase,
                **checklist_scope_attributes(sheet, category, check, phase),
            }
            items.append(item)
    wb.close()
    ids = [item["id"] for item in items]
    if len(ids) != len(set(ids)):
        raise GovernanceError("checklist contains duplicate IDs")
    return {
        "schema_version": 2,
        "generated_at": utcnow(),
        "source": str(book_path.relative_to(ROOT)),
        "source_sha256": sha256_file(book_path),
        "item_count": len(items),
        "items": items,
    }


def stable_catalog(catalog: dict[str, Any]) -> dict[str, Any]:
    result = dict(catalog)
    result.pop("generated_at", None)
    return result


def cmd_catalog(args: argparse.Namespace) -> int:
    candidate = workbook_catalog()
    if args.check:
        current = load_catalog()
        if stable_catalog(candidate) != stable_catalog(current):
            print("catalog is stale; run: python tools/devflow.py catalog", file=sys.stderr)
            return 1
        print(f"catalog OK: {current['item_count']} items / {current['source_sha256']}")
        return 0
    atomic_write_json(CATALOG_PATH, candidate)
    print(f"wrote {CATALOG_PATH.relative_to(ROOT)}: {candidate['item_count']} items")
    return 0


def chained_records(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    records: list[dict[str, Any]] = []
    for line_number, line in enumerate(path.read_text(encoding="utf-8").splitlines(), 1):
        if not line.strip():
            continue
        try:
            records.append(json.loads(line))
        except json.JSONDecodeError as exc:
            raise GovernanceError(f"{path}: invalid JSONL at line {line_number}") from exc
    return records


def verify_chain(path: Path) -> list[dict[str, Any]]:
    previous = "GENESIS"
    records = chained_records(path)
    for index, record in enumerate(records, 1):
        if record.get("index") != index:
            raise GovernanceError(f"{path}: record index mismatch at {index}")
        if record.get("previous_hash") != previous:
            raise GovernanceError(f"{path}: previous hash mismatch at {index}")
        payload = dict(record)
        actual = payload.pop("record_hash", None)
        expected = sha256_bytes(canonical_bytes(payload))
        if actual != expected:
            raise GovernanceError(f"{path}: record hash mismatch at {index}")
        previous = str(actual)
    return records


def append_chained(path: Path, payload: dict[str, Any]) -> dict[str, Any]:
    records = verify_chain(path)
    record = {
        "index": len(records) + 1,
        "previous_hash": records[-1]["record_hash"] if records else "GENESIS",
        **payload,
    }
    record["record_hash"] = sha256_bytes(canonical_bytes(record))
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as stream:
        stream.write(json.dumps(record, ensure_ascii=False, sort_keys=True) + "\n")
        stream.flush()
        os.fsync(stream.fileno())
    return record


def item_profiles(policy: dict[str, Any], profiles: list[str]) -> set[str]:
    sheets: set[str] = set()
    for profile in profiles:
        if profile not in policy["profiles"]:
            raise GovernanceError(f"unknown profile: {profile}")
        sheets.update(policy["profiles"][profile]["sheets"])
    return sheets


def normalize_profiles(values: Iterable[str]) -> list[str]:
    profiles: list[str] = []
    for value in values:
        for profile in value.split(","):
            profile = profile.strip().upper()
            if profile and profile not in profiles:
                profiles.append(profile)
    if "CORE" not in profiles:
        profiles.insert(0, "CORE")
    if any(profile.endswith("-DELTA") for profile in profiles) and "CLOUD-COMMON" not in profiles:
        profiles.insert(1, "CLOUD-COMMON")
    return profiles


def result_history_snapshot(item: dict[str, Any]) -> dict[str, Any]:
    """現在のレビュー判断を、再確認可能な履歴レコードへ変換する。"""
    fields = [
        "applicability",
        "na_rationale",
        "verdict",
        "evidence",
        "issue_id",
        "reviewer",
        "reviewed_at",
        "project_severity",
        "severity_rationale",
        "remediation_plan",
        "due_at",
        "recheck",
    ]
    return {field: item.get(field) for field in fields}


def append_result_history(item: dict[str, Any]) -> None:
    if item.get("applicability") == "undecided" and item.get("verdict") == "unreviewed":
        return
    history = item.setdefault("history", [])
    if not isinstance(history, list):
        raise GovernanceError(f"{item.get('id')}: invalid review history")
    history.append(result_history_snapshot(item))


def parse_due_at(value: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise GovernanceError("是正期限はYYYY-MM-DD形式で指定する") from exc


def recheck_valid(work: Path, recheck: Any) -> tuple[bool, str]:
    if not isinstance(recheck, dict):
        return False, "Fail是正後の再確認記録がない"
    required = ["verdict", "evidence", "reviewer", "reviewed_at"]
    missing = [field for field in required if not recheck.get(field)]
    if missing:
        return False, f"再確認項目が不足: {', '.join(missing)}"
    if recheck["verdict"] != "pass":
        return False, "再確認結果がPassでない"
    evidence = recheck["evidence"]
    if not isinstance(evidence, list) or not evidence_valid(work, [str(value) for value in evidence]):
        return False, "再確認を裏付ける到達可能な証跡がない"
    return True, ""


def cmd_init(args: argparse.Namespace) -> int:
    policy = load_policy()
    catalog = load_catalog()
    profiles = normalize_profiles(args.profile)
    sheets = item_profiles(policy, profiles)
    work_id = args.id or f"WI-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:6]}"
    if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._-]{2,120}", work_id):
        raise GovernanceError("invalid work item ID")
    work = WORK_ROOT / work_id
    if work.exists():
        raise GovernanceError(f"work item already exists: {work_id}")
    scope_state = None
    selected_ids = None
    if getattr(args, "scope_file", None):
        scope_path = Path(args.scope_file)
        if not scope_path.is_file():
            raise GovernanceError(f"execution scope not found: {scope_path}")
        scopeflow = load_scopeflow()
        scope_state = read_json(scope_path)
        scope_errors = scopeflow.validate_scope(scope_state)
        if scope_errors:
            raise GovernanceError("invalid execution scope: " + "; ".join(scope_errors))
        selection = scopeflow.select_catalog_items(
            catalog,
            scope_state,
            profiles=profiles,
            phases=list(policy["phases"]),
        )
        if not selection["selected_item_ids"]:
            raise GovernanceError("execution scope selected no checklist items")
        scope_state["selection"] = selection
        scope_state["updated_at"] = utcnow()
        selected_ids = set(selection["selected_item_ids"])
    selected = []
    for item in catalog["items"]:
        if item["sheet"] not in sheets:
            continue
        if selected_ids is not None and item["id"] not in selected_ids:
            continue
        selected.append({
            "id": item["id"],
            "sheet": item["sheet"],
            "phase": item["phase"],
            "base_severity": item["base_severity"],
            "project_severity": item["base_severity"],
            "severity_rationale": "",
            "applicability": "undecided",
            "na_rationale": "",
            "verdict": "unreviewed",
            "evidence": [],
            "issue_id": "",
            "remediation_plan": "",
            "due_at": "",
            "reviewer": "",
            "reviewed_at": "",
            "recheck": None,
            "history": [],
            "exception": None,
        })
    for directory in [work / "docs", work / "review", work / "evidence", work / "reports"]:
        directory.mkdir(parents=True, exist_ok=True)
    created = utcnow()
    replacements = {
        "{{WORK_ITEM}}": work_id,
        "{{TITLE}}": args.title,
        "{{CREATED_AT}}": created,
        "{{PROFILES}}": ", ".join(profiles),
        "{{REQUEST}}": args.request,
    }
    for template in sorted(TEMPLATE_ROOT.glob("*.md")):
        content = template.read_text(encoding="utf-8")
        for old, new in replacements.items():
            content = content.replace(old, new)
        (work / "docs" / template.name).write_text(content, encoding="utf-8")
    if scope_state:
        atomic_write_json(work / "execution-scope.json", scope_state)
    state = {
        "schema_version": 1,
        "workflow_schema_version": policy["schema_version"],
        "id": work_id,
        "title": args.title,
        "created_at": created,
        "updated_at": created,
        "current_phase": "intake",
        "status": "active",
        "profiles": profiles,
        "catalog_sha256": sha256_bytes(canonical_bytes(stable_catalog(catalog))),
        "execution_scope_sha256": sha256_bytes(canonical_bytes(scope_state)) if scope_state else "",
    }
    atomic_write_json(work / "state.json", state)
    atomic_write_json(
        work / "review" / "checklist-results.json",
        {"schema_version": RESULTS_SCHEMA_VERSION, "items": selected},
    )
    (work / "approvals.jsonl").touch()
    (work / "events.jsonl").touch()
    append_chained(work / "events.jsonl", {
        "timestamp": created,
        "event": "work-item-created",
        "phase": "intake",
        "actor": args.actor,
        "details": {
            "profiles": profiles,
            "catalog_items": len(selected),
            "selector_version": scope_state.get("selection", {}).get("selector_version", "legacy-full-profile") if scope_state else "legacy-full-profile",
        },
    })
    try:
        display_path = work.relative_to(ROOT)
    except ValueError:
        display_path = work
    print(work_id)
    print(f"created {display_path} with {len(selected)} checklist items")
    return 0


def load_work(work_item: str) -> tuple[Path, dict[str, Any], dict[str, Any]]:
    work = safe_work_path(work_item)
    results = read_json(work / "review" / "checklist-results.json")
    validate_results_envelope(results)
    return work, read_json(work / "state.json"), results


def validate_results_envelope(results: Any) -> None:
    if not isinstance(results, dict) or set(results) != {"schema_version", "items"}:
        raise GovernanceError("invalid checklist result envelope")
    if results["schema_version"] not in {1, RESULTS_SCHEMA_VERSION} or not isinstance(results["items"], list):
        raise GovernanceError("unsupported checklist result schema")
    if results["schema_version"] < RESULTS_SCHEMA_VERSION:
        return
    required = {
        "severity_rationale",
        "remediation_plan",
        "due_at",
        "recheck",
        "history",
    }
    for item in results["items"]:
        if not isinstance(item, dict) or not required <= set(item):
            raise GovernanceError(f"{item.get('id') if isinstance(item, dict) else '?'}: auditable result fields missing")
        if not isinstance(item["history"], list):
            raise GovernanceError(f"{item['id']}: review history must be a list")


def require_current_workflow(state: dict[str, Any]) -> None:
    current = load_policy()["schema_version"]
    actual = state.get("workflow_schema_version")
    if actual != current:
        shown = "legacy" if actual is None else str(actual)
        raise GovernanceError(
            f"work item workflow schema is {shown}; run migrate --work-item {state['id']} before continuing"
        )


def authorization_definition() -> tuple[str, str]:
    authorization = load_policy().get("authorization")
    if not isinstance(authorization, dict):
        raise GovernanceError("policy authorization definition is missing")
    phase = str(authorization.get("phase") or "")
    role = str(authorization.get("role") or "")
    if phase not in load_policy()["phases"] or not role:
        raise GovernanceError("policy authorization definition is invalid")
    return phase, role


def evidence_valid(work: Path, evidence: list[str]) -> bool:
    for value in evidence:
        if re.match(r"https?://", value):
            return True
        candidate = Path(value)
        if not candidate.is_absolute():
            for base in [work, ROOT]:
                if (base / candidate).exists():
                    return True
        elif candidate.exists():
            return True
    return False


def required_doc_state(work: Path, relative: str, placeholders: list[str]) -> tuple[dict[str, str] | None, list[dict[str, str]]]:
    path = work / relative
    blockers: list[dict[str, str]] = []
    if not path.is_file():
        blockers.append({"code": "DOC_MISSING", "subject": relative, "message": "必須文書が存在しない"})
        return None, blockers
    content = path.read_text(encoding="utf-8")
    if len(content.strip()) < 120:
        blockers.append({"code": "DOC_TOO_SHORT", "subject": relative, "message": "必須文書が実質的な内容を持たない"})
    found = [token for token in placeholders if token in content]
    if found:
        blockers.append({"code": "DOC_PLACEHOLDER", "subject": relative, "message": f"未記入トークンが残る: {', '.join(found)}"})
    return {"path": relative, "sha256": sha256_file(path)}, blockers


def exception_valid(item: dict[str, Any]) -> tuple[bool, str]:
    exception = item.get("exception")
    if not isinstance(exception, dict) or exception.get("status") != "approved":
        return False, "Failに承認済み例外がない"
    required = ["approver", "role", "rationale", "expires_at"]
    missing = [field for field in required if not exception.get(field)]
    if missing:
        return False, f"例外項目が不足: {', '.join(missing)}"
    try:
        expires = datetime.fromisoformat(str(exception["expires_at"]).replace("Z", "+00:00"))
    except ValueError:
        return False, "例外期限がISO 8601形式でない"
    if expires <= datetime.now(timezone.utc):
        return False, "例外承認が期限切れ"
    if item.get("project_severity") == "Critical" and exception.get("role") not in {"security-owner", "release-owner", "governance-owner"}:
        return False, "案件Criticalの例外承認者ロールが不足"
    return True, ""


def gate_snapshot(work: Path, state: dict[str, Any], results: dict[str, Any], phase: str, *, include_approvals: bool) -> dict[str, Any]:
    policy = load_policy()
    if phase not in policy["phases"]:
        raise GovernanceError(f"unknown phase: {phase}")
    definition = policy["phases"][phase]
    blockers: list[dict[str, str]] = []
    documents: list[dict[str, str]] = []
    for relative in definition["required_docs"]:
        document, failures = required_doc_state(work, relative, policy["document_placeholders"])
        if document:
            documents.append(document)
        blockers.extend(failures)
    review_slice = sorted((item for item in results["items"] if item["phase"] == phase), key=lambda item: item["id"])
    auditable_results = int(results.get("schema_version", 1)) >= RESULTS_SCHEMA_VERSION
    for item in review_slice:
        item_id = item["id"]
        applicability = item.get("applicability")
        if applicability == "undecided":
            blockers.append({"code": "CHECK_APPLICABILITY", "subject": item_id, "message": "適用判定が未決定"})
            continue
        if not str(item.get("reviewer") or "").strip() or not str(item.get("reviewed_at") or "").strip():
            blockers.append({"code": "CHECK_REVIEWER", "subject": item_id, "message": "レビュアーまたはレビュー日時がない"})
        if auditable_results and not str(item.get("severity_rationale") or "").strip():
            blockers.append({"code": "CHECK_SEVERITY_RATIONALE", "subject": item_id, "message": "案件重要度の判断根拠がない"})
        if applicability == "not-applicable":
            if not str(item.get("na_rationale") or "").strip():
                blockers.append({"code": "CHECK_NA_RATIONALE", "subject": item_id, "message": "N/A根拠がない"})
            continue
        if applicability != "applicable":
            blockers.append({"code": "CHECK_VALUE", "subject": item_id, "message": f"不正な適用判定: {applicability}"})
            continue
        verdict = item.get("verdict")
        if verdict == "pass":
            evidence = item.get("evidence") or []
            if not isinstance(evidence, list) or not evidence_valid(work, [str(value) for value in evidence]):
                blockers.append({"code": "CHECK_EVIDENCE", "subject": item_id, "message": "Passを裏付ける到達可能な証跡がない"})
            if auditable_results and any(
                isinstance(entry, dict) and entry.get("verdict") == "fail"
                for entry in item.get("history", [])
            ):
                valid, reason = recheck_valid(work, item.get("recheck"))
                if not valid:
                    blockers.append({"code": "CHECK_RECHECK", "subject": item_id, "message": reason})
        elif verdict == "fail":
            if not str(item.get("issue_id") or "").strip():
                blockers.append({"code": "CHECK_ISSUE", "subject": item_id, "message": "FailにIssue IDがない"})
            if auditable_results:
                if not str(item.get("remediation_plan") or "").strip():
                    blockers.append({"code": "CHECK_REMEDIATION", "subject": item_id, "message": "Failに対応方針がない"})
                due_at = str(item.get("due_at") or "")
                try:
                    if not due_at:
                        raise ValueError
                    due = date.fromisoformat(due_at)
                    if due < datetime.now(timezone.utc).date():
                        blockers.append({"code": "CHECK_DUE_OVERDUE", "subject": item_id, "message": "是正期限を超過している"})
                except ValueError:
                    blockers.append({"code": "CHECK_DUE", "subject": item_id, "message": "Failに有効な是正期限がない"})
            if state.get("workflow_schema_version") == policy["schema_version"]:
                blockers.append({
                    "code": "CHECK_FAIL_BLOCKING",
                    "subject": item_id,
                    "message": "単一承認workflowではFailを後続承認で受容できない。原因を修正すること",
                })
            else:
                valid, reason = exception_valid(item)
                if not valid:
                    blockers.append({"code": "CHECK_EXCEPTION", "subject": item_id, "message": reason})
        else:
            blockers.append({"code": "CHECK_VERDICT", "subject": item_id, "message": "適用項目の判定が未確認"})
    digest_payload = {
        "work_item": state["id"],
        "phase": phase,
        "documents": documents,
        "reviews": review_slice,
    }
    gate_digest = sha256_bytes(canonical_bytes(digest_payload))
    approval_status: dict[str, dict[str, Any] | None] = {role: None for role in definition["required_approvals"]}
    if include_approvals:
        approvals = verify_chain(work / "approvals.jsonl")
        for record in approvals:
            if record.get("phase") == phase and record.get("gate_digest") == gate_digest and record.get("role") in approval_status:
                approval_status[record["role"]] = record
        for role, record in approval_status.items():
            if record is None:
                blockers.append({"code": "APPROVAL_MISSING", "subject": role, "message": "現行成果物に対する承認がない"})
            elif record.get("decision") != "approved":
                blockers.append({"code": "APPROVAL_REJECTED", "subject": role, "message": "最新の承認判断がapprovedでない"})
    return {
        "schema_version": 1,
        "work_item": state["id"],
        "phase": phase,
        "generated_at": utcnow(),
        "gate_digest": gate_digest,
        "documents": documents,
        "review_item_count": len(review_slice),
        "approval_status": approval_status,
        "blockers": blockers,
        "passed": not blockers,
    }


def write_inspection(work: Path, report: dict[str, Any]) -> None:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    atomic_write_json(work / "reports" / f"{stamp}-{report['phase']}.json", report)
    atomic_write_json(work / "reports" / f"latest-{report['phase']}.json", report)


def print_report(report: dict[str, Any], as_json: bool) -> None:
    if as_json:
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return
    status = "PASS" if report["passed"] else "BLOCKED"
    print(f"{status}: {report['work_item']} / {report['phase']} / digest {report['gate_digest']}")
    print(f"documents={len(report['documents'])} reviews={report['review_item_count']} blockers={len(report['blockers'])}")
    for blocker in report["blockers"]:
        print(f"- [{blocker['code']}] {blocker['subject']}: {blocker['message']}")


def cmd_inspect(args: argparse.Namespace) -> int:
    work, state, results = load_work(args.work_item)
    phase = args.phase or state["current_phase"]
    report = gate_snapshot(work, state, results, phase, include_approvals=not args.ignore_approvals)
    write_inspection(work, report)
    print_report(report, args.json)
    return 0 if report["passed"] else 1


def cmd_set_check(args: argparse.Namespace) -> int:
    policy = load_policy()
    work, state, results = load_work(args.work_item)
    match = next((item for item in results["items"] if item["id"] == args.item), None)
    if match is None:
        raise GovernanceError(f"checklist item not selected by profiles: {args.item}")
    if args.applicability not in policy["review_values"]["applicability"]:
        raise GovernanceError("invalid applicability")
    if args.verdict not in policy["review_values"]["verdict"]:
        raise GovernanceError("invalid verdict")
    if args.severity not in policy["review_values"]["severity"]:
        raise GovernanceError("invalid project severity")
    auditable_results = int(results.get("schema_version", 1)) >= RESULTS_SCHEMA_VERSION
    if args.applicability == "not-applicable" and not str(args.na_rationale or "").strip():
        raise GovernanceError("N/A rationale required")
    if args.applicability == "applicable" and args.verdict == "pass" and not args.evidence:
        raise GovernanceError("Pass evidence required")
    if args.applicability == "applicable" and args.verdict == "fail" and not str(args.issue or "").strip():
        raise GovernanceError("Fail issue required")
    severity_rationale = str(getattr(args, "severity_rationale", "") or "")
    remediation_plan = str(getattr(args, "remediation_plan", "") or "")
    due_at = str(getattr(args, "due_at", "") or "")
    recheck_evidence = getattr(args, "recheck_evidence", None)
    recheck_reviewer = str(getattr(args, "recheck_reviewer", "") or "")
    if auditable_results and not severity_rationale.strip():
        raise GovernanceError("project severity rationale required")
    if auditable_results and args.applicability == "applicable" and args.verdict == "fail":
        if not remediation_plan.strip():
            raise GovernanceError("Fail remediation plan required")
        parse_due_at(due_at)
    if args.exception_approver and state.get("workflow_schema_version") == policy["schema_version"]:
        raise GovernanceError("single-authorization workflow does not accept later human risk exceptions")
    append_result_history(match)
    reviewed_at = utcnow()
    recheck = None
    if recheck_evidence or recheck_reviewer:
        recheck = {
            "verdict": "pass",
            "evidence": recheck_evidence or [],
            "reviewer": recheck_reviewer or args.reviewer,
            "reviewed_at": reviewed_at,
        }
    match.update({
        "applicability": args.applicability,
        "na_rationale": args.na_rationale or "",
        "verdict": args.verdict,
        "evidence": args.evidence or [],
        "issue_id": args.issue or "",
        "remediation_plan": remediation_plan,
        "due_at": due_at,
        "reviewer": args.reviewer,
        "reviewed_at": reviewed_at,
        "project_severity": args.severity,
        "severity_rationale": severity_rationale,
        "recheck": recheck,
    })
    if args.exception_approver:
        match["exception"] = {
            "status": "approved",
            "approver": args.exception_approver,
            "role": args.exception_role,
            "rationale": args.exception_rationale,
            "expires_at": args.exception_expires,
        }
    elif args.clear_exception:
        match["exception"] = None
    atomic_write_json(work / "review" / "checklist-results.json", results)
    state["updated_at"] = reviewed_at
    atomic_write_json(work / "state.json", state)
    append_chained(work / "events.jsonl", {
        "timestamp": reviewed_at,
        "event": "checklist-result-updated",
        "phase": match["phase"],
        "actor": args.reviewer,
        "details": {"item": args.item, "applicability": args.applicability, "verdict": args.verdict},
    })
    print(f"updated {args.item}")
    return 0


def cmd_set_checks(args: argparse.Namespace) -> int:
    """Atomically record a validated batch of checklist dispositions."""
    policy = load_policy()
    work, state, results = load_work(args.work_item)
    payload = read_json(Path(args.input))
    if not isinstance(payload, list) or not payload:
        raise GovernanceError("batch input must be a non-empty JSON list")
    by_id = {item["id"]: item for item in results["items"]}
    auditable_results = int(results.get("schema_version", 1)) >= RESULTS_SCHEMA_VERSION
    seen: set[str] = set()
    validated: list[tuple[dict[str, Any], dict[str, Any]]] = []
    for entry in payload:
        if not isinstance(entry, dict):
            raise GovernanceError("batch entry must be an object")
        item_id = str(entry.get("item") or "")
        if item_id in seen:
            raise GovernanceError(f"duplicate batch item: {item_id}")
        seen.add(item_id)
        match = by_id.get(item_id)
        if match is None:
            raise GovernanceError(f"checklist item not selected by profiles: {item_id}")
        applicability = entry.get("applicability")
        verdict = entry.get("verdict")
        severity = entry.get("severity")
        if applicability not in policy["review_values"]["applicability"]:
            raise GovernanceError(f"{item_id}: invalid applicability")
        if verdict not in policy["review_values"]["verdict"]:
            raise GovernanceError(f"{item_id}: invalid verdict")
        if severity not in policy["review_values"]["severity"]:
            raise GovernanceError(f"{item_id}: invalid project severity")
        if applicability == "not-applicable" and not str(entry.get("na_rationale") or "").strip():
            raise GovernanceError(f"{item_id}: N/A rationale required")
        if applicability == "applicable" and verdict == "pass" and not entry.get("evidence"):
            raise GovernanceError(f"{item_id}: Pass evidence required")
        if applicability == "applicable" and verdict == "fail" and not str(entry.get("issue") or "").strip():
            raise GovernanceError(f"{item_id}: Fail issue required")
        if not str(entry.get("reviewer") or "").strip():
            raise GovernanceError(f"{item_id}: reviewer required")
        if auditable_results and not str(entry.get("severity_rationale") or "").strip():
            raise GovernanceError(f"{item_id}: project severity rationale required")
        if auditable_results and applicability == "applicable" and verdict == "fail":
            if not str(entry.get("remediation_plan") or "").strip():
                raise GovernanceError(f"{item_id}: Fail remediation plan required")
            try:
                parse_due_at(str(entry.get("due_at") or ""))
            except GovernanceError as exc:
                raise GovernanceError(f"{item_id}: {exc}") from exc
        failed_before = match.get("verdict") == "fail" or any(
            isinstance(history, dict) and history.get("verdict") == "fail"
            for history in match.get("history", [])
        )
        if auditable_results and applicability == "applicable" and verdict == "pass" and failed_before:
            valid, reason = recheck_valid(work, entry.get("recheck"))
            if not valid:
                raise GovernanceError(f"{item_id}: {reason}")
        validated.append((match, entry))
    reviewed_at = utcnow()
    for match, entry in validated:
        append_result_history(match)
        match.update({
            "applicability": entry["applicability"],
            "na_rationale": str(entry.get("na_rationale") or ""),
            "verdict": entry["verdict"],
            "evidence": [str(value) for value in entry.get("evidence") or []],
            "issue_id": str(entry.get("issue") or ""),
            "remediation_plan": str(entry.get("remediation_plan") or ""),
            "due_at": str(entry.get("due_at") or ""),
            "reviewer": str(entry["reviewer"]),
            "reviewed_at": reviewed_at,
            "project_severity": entry["severity"],
            "severity_rationale": str(entry.get("severity_rationale") or ""),
            "recheck": entry.get("recheck"),
            "exception": None,
        })
    atomic_write_json(work / "review" / "checklist-results.json", results)
    state["updated_at"] = reviewed_at
    atomic_write_json(work / "state.json", state)
    append_chained(work / "events.jsonl", {
        "timestamp": reviewed_at,
        "event": "checklist-batch-updated",
        "phase": state["current_phase"],
        "actor": args.actor,
        "details": {
            "count": len(validated),
            "items_sha256": sha256_bytes(canonical_bytes(sorted(seen))),
        },
    })
    print(f"updated {len(validated)} checklist items atomically")
    return 0


def cmd_approve(args: argparse.Namespace) -> int:
    work, state, results = load_work(args.work_item)
    require_current_workflow(state)
    authorization_phase, authorization_role = authorization_definition()
    phase = getattr(args, "phase", None) or authorization_phase
    if state["current_phase"] != authorization_phase:
        raise GovernanceError(
            f"initial authorization can be recorded only while current phase is {authorization_phase}"
        )
    policy = load_policy()
    allowed_roles = policy["phases"][phase]["required_approvals"]
    role = getattr(args, "role", None) or authorization_role
    if phase != authorization_phase or role != authorization_role:
        raise GovernanceError(
            f"schema {policy['schema_version']} permits human authorization only for "
            f"{authorization_phase} / {authorization_role}"
        )
    if role not in allowed_roles:
        raise GovernanceError(f"role {role} is not an authorization role for {phase}")
    report = gate_snapshot(work, state, results, phase, include_approvals=False)
    content_blockers = [blocker for blocker in report["blockers"] if not blocker["code"].startswith("APPROVAL_")]
    if args.decision == "approved" and content_blockers:
        print_report(report, False)
        raise GovernanceError("cannot approve a gate with unresolved content/checklist blockers")
    record = append_chained(work / "approvals.jsonl", {
        "timestamp": utcnow(),
        "work_item": state["id"],
        "phase": phase,
        "gate_digest": report["gate_digest"],
        "decision": args.decision,
        "approver": args.approver,
        "role": role,
        "comment": args.comment,
    })
    append_chained(work / "events.jsonl", {
        "timestamp": utcnow(),
        "event": "initial-authorization-recorded",
        "phase": phase,
        "actor": args.approver,
        "details": {"role": role, "decision": args.decision, "approval_hash": record["record_hash"]},
    })
    print(f"recorded {args.decision}: {phase} / {role} / {record['record_hash']}")
    return 0


def cmd_migrate(args: argparse.Namespace) -> int:
    work, state, _ = load_work(args.work_item)
    target = load_policy()["schema_version"]
    if state.get("workflow_schema_version") == target:
        raise GovernanceError(f"work item already uses workflow schema {target}")
    if state.get("current_phase") not in {"intake", "requirements"}:
        raise GovernanceError(
            "legacy work items beyond requirements cannot be migrated in place; create a new work item"
        )
    execution_plan = work / "docs" / "01-execution-plan.md"
    if not execution_plan.is_file():
        raise GovernanceError("migration requires docs/01-execution-plan.md")
    previous = state.get("workflow_schema_version")
    state["workflow_schema_version"] = target
    state["updated_at"] = utcnow()
    atomic_write_json(work / "state.json", state)
    append_chained(work / "events.jsonl", {
        "timestamp": utcnow(),
        "event": "workflow-migrated",
        "phase": state["current_phase"],
        "actor": args.actor,
        "details": {"from": previous, "to": target, "prior_authorizations_reused": False},
    })
    print(f"migrated {state['id']}: {previous or 'legacy'} -> {target}; fresh authorization required")
    return 0


def preceding_phase_reports(
    work: Path,
    state: dict[str, Any],
    results: dict[str, Any],
    phase: str,
) -> list[dict[str, Any]]:
    policy = load_policy()
    order = policy["phase_order"]
    position = order.index(phase)
    return [
        gate_snapshot(work, state, results, previous, include_approvals=True)
        for previous in order[:position]
    ]


def cmd_advance(args: argparse.Namespace) -> int:
    work, state, results = load_work(args.work_item)
    require_current_workflow(state)
    current = state["current_phase"]
    prior_reports = preceding_phase_reports(work, state, results, current)
    failed_prior = next((report for report in prior_reports if not report["passed"]), None)
    if failed_prior is not None:
        write_inspection(work, failed_prior)
        print_report(failed_prior, False)
        raise GovernanceError(f"preceding phase gate is no longer valid: {failed_prior['phase']}")
    report = gate_snapshot(work, state, results, current, include_approvals=True)
    write_inspection(work, report)
    if not report["passed"]:
        print_report(report, False)
        raise GovernanceError("current phase gate is blocked")
    order = load_policy()["phase_order"]
    position = order.index(current)
    if position == len(order) - 1:
        raise GovernanceError("work item is already closed")
    next_phase = order[position + 1]
    state["current_phase"] = next_phase
    state["updated_at"] = utcnow()
    if next_phase == "closed":
        state["status"] = "closed"
    atomic_write_json(work / "state.json", state)
    append_chained(work / "events.jsonl", {
        "timestamp": utcnow(),
        "event": "phase-advanced",
        "phase": next_phase,
        "actor": args.actor,
        "details": {"from": current, "to": next_phase, "gate_digest": report["gate_digest"]},
    })
    print(f"advanced {state['id']}: {current} -> {next_phase}")
    return 0


def cmd_status(args: argparse.Namespace) -> int:
    work, state, results = load_work(args.work_item)
    require_current_workflow(state)
    report = gate_snapshot(work, state, results, state["current_phase"], include_approvals=True)
    authorization_phase, _ = authorization_definition()
    authorization = gate_snapshot(work, state, results, authorization_phase, include_approvals=True)
    payload = {**state, "gate": report, "initial_authorization": authorization}
    if args.json:
        print(json.dumps(payload, ensure_ascii=False, indent=2))
    else:
        print(f"{state['id']}: {state['title']}")
        print(f"phase={state['current_phase']} status={state['status']} profiles={','.join(state['profiles'])}")
        print(f"gate={'PASS' if report['passed'] else 'BLOCKED'} blockers={len(report['blockers'])}")
        print(f"initial_authorization={'VALID' if authorization['passed'] else 'MISSING_OR_STALE'}")
    return 0


def load_improvements() -> list[dict[str, Any]]:
    if not IMPROVEMENTS_PATH.exists():
        return []
    value = read_json(IMPROVEMENTS_PATH)
    if not isinstance(value, list):
        raise GovernanceError("improvement proposals must be a JSON list")
    return value


def save_improvements(proposals: list[dict[str, Any]]) -> None:
    atomic_write_json(IMPROVEMENTS_PATH, proposals)


def propose_improvement(skill: str, problem: str, change: str, evidence: list[str], session_id: str) -> dict[str, Any]:
    proposals = load_improvements()
    proposal_id = "IMP-" + sha256_bytes(canonical_bytes({"skill": skill, "problem": problem, "change": change}))[:12]
    existing = next((item for item in proposals if item["id"] == proposal_id), None)
    if existing:
        for value in evidence:
            if value not in existing["evidence"]:
                existing["evidence"].append(value)
        existing["last_seen_at"] = utcnow()
        existing["sessions"] = sorted(set(existing.get("sessions", []) + [session_id]))
        save_improvements(proposals)
        return existing
    proposal = {
        "id": proposal_id,
        "skill": skill,
        "problem": problem,
        "change": change,
        "evidence": evidence,
        "sessions": [session_id],
        "created_at": utcnow(),
        "last_seen_at": utcnow(),
        "status": "pending",
        "approved_by": "",
        "approved_at": "",
        "applied_at": "",
    }
    proposals.append(proposal)
    save_improvements(proposals)
    return proposal


def skill_for_blocker(code: str) -> str:
    if code.startswith("DOC_"):
        return "author-lifecycle-docs"
    if code.startswith("CHECK_"):
        return "inspect-quality-gates"
    if code.startswith("APPROVAL_"):
        return "authorize-autonomous-execution"
    return "govern-development-request"


def session_retrospective(session_id: str, cwd: str) -> Path:
    REPORT_ROOT.mkdir(parents=True, exist_ok=True)
    retro_dir = REPORT_ROOT / "retrospectives"
    retro_dir.mkdir(parents=True, exist_ok=True)
    active: list[tuple[Path, dict[str, Any], dict[str, Any], dict[str, Any]]] = []
    failure_log = REPORT_ROOT / "gate-failures.jsonl"
    for state_path in sorted(WORK_ROOT.glob("*/state.json")) if WORK_ROOT.exists() else []:
        state = read_json(state_path)
        if state.get("status") != "active":
            continue
        work = state_path.parent
        results = read_json(work / "review" / "checklist-results.json")
        report = gate_snapshot(work, state, results, state["current_phase"], include_approvals=True)
        active.append((work, state, results, report))
        for blocker in report["blockers"]:
            failure_log.parent.mkdir(parents=True, exist_ok=True)
            with failure_log.open("a", encoding="utf-8") as stream:
                stream.write(json.dumps({"timestamp": utcnow(), "session_id": session_id, "work_item": state["id"], "phase": state["current_phase"], **blocker}, ensure_ascii=False, sort_keys=True) + "\n")
    git_status = ""
    try:
        git_status = subprocess.run(["git", "status", "--short"], cwd=ROOT, text=True, capture_output=True, check=False, timeout=10).stdout.strip()
    except (OSError, subprocess.SubprocessError):
        git_status = "取得失敗"
    lines = [
        "# セッション振り返り",
        "",
        f"- Session: {session_id}",
        f"- Generated: {utcnow()}",
        f"- CWD: {cwd}",
        "",
        "## 作業ツリー",
        "",
        "```text",
        git_status or "clean",
        "```",
        "",
        "## アクティブ作業とゲート",
        "",
    ]
    if not active:
        lines.append("- アクティブなwork itemなし")
    for _, state, _, report in active:
        lines.append(f"- {state['id']} / {state['current_phase']}: {'PASS' if report['passed'] else 'BLOCKED'} ({len(report['blockers'])} blockers)")
        for blocker in report["blockers"][:20]:
            lines.append(f"  - [{blocker['code']}] {blocker['subject']}: {blocker['message']}")
        if len(report["blockers"]) > 20:
            lines.append(f"  - ほか {len(report['blockers']) - 20} 件")
    lines.extend(["", "## 改善候補", ""])
    failures = []
    if failure_log.exists():
        for line in failure_log.read_text(encoding="utf-8").splitlines():
            if line.strip():
                failures.append(json.loads(line))
    generated: list[dict[str, Any]] = []
    for code in sorted({entry["code"] for entry in failures}):
        matching = [entry for entry in failures if entry["code"] == code]
        sessions = {entry["session_id"] for entry in matching}
        if len(sessions) < 2:
            continue
        latest = matching[-1]
        generated.append(propose_improvement(
            skill_for_blocker(code),
            f"{code}が複数セッションで再発: {latest['message']}",
            f"{code}を作業開始時に予防し、ゲート前に検出する具体的手順を追加する。",
            [f"{entry['work_item']}:{entry['phase']}:{entry['subject']}" for entry in matching[-5:]],
            session_id,
        ))
    if generated:
        for proposal in generated:
            lines.append(f"- {proposal['id']} -> {proposal['skill']}: {proposal['problem']}（状態: {proposal['status']}）")
    else:
        lines.append("- 再発条件を満たす自動改善候補なし")
    lines.extend([
        "",
        "## 改善候補の扱い",
        "",
        "改善候補は自動適用しない。現在の実行計画外であれば別work itemへ分離し、要件と実行計画の初回承認後に実装する。",
        "",
    ])
    filename = f"{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}-{re.sub(r'[^A-Za-z0-9._-]', '_', session_id)[:48]}.md"
    path = retro_dir / filename
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def cmd_session_retrospective(args: argparse.Namespace) -> int:
    path = session_retrospective(args.session_id, args.cwd)
    print(path.relative_to(ROOT))
    return 0


def cmd_improvement_list(args: argparse.Namespace) -> int:
    proposals = load_improvements()
    selected = [item for item in proposals if not args.status or item["status"] == args.status]
    if args.json:
        print(json.dumps(selected, ensure_ascii=False, indent=2))
    else:
        for item in selected:
            print(f"{item['id']} [{item['status']}] {item['skill']}: {item['problem']}")
    return 0


def cmd_audit(args: argparse.Namespace) -> int:
    failures: list[str] = []
    warnings: list[str] = []
    try:
        candidate = workbook_catalog()
        current = load_catalog()
        if stable_catalog(candidate) != stable_catalog(current):
            failures.append("catalog is stale")
    except GovernanceError as exc:
        failures.append(str(exc))
    if WORK_ROOT.exists():
        for state_path in sorted(WORK_ROOT.glob("*/state.json")):
            work = state_path.parent
            try:
                state = read_json(state_path)
                results = read_json(work / "review" / "checklist-results.json")
                validate_results_envelope(results)
                verify_chain(work / "events.jsonl")
                verify_chain(work / "approvals.jsonl")
                require_current_workflow(state)
                ids = [item["id"] for item in results["items"]]
                if len(ids) != len(set(ids)):
                    failures.append(f"{state['id']}: duplicate checklist results")
                scope_path = work / "execution-scope.json"
                if scope_path.is_file():
                    scopeflow = load_scopeflow()
                    scope_report = scopeflow.audit_scope(read_json(scope_path))
                    failures.extend(f"{state['id']}: execution scope: {message}" for message in scope_report["errors"])
                    warnings.extend(f"{state['id']}: execution scope: {message}" for message in scope_report["warnings"])
                    if state.get("status") == "closed" and not (work / "reports" / "execution-efficiency.json").is_file():
                        failures.append(f"{state['id']}: execution efficiency report missing")
                if state["current_phase"] not in load_policy()["phase_order"]:
                    failures.append(f"{state['id']}: invalid phase")
                    continue
                phases_to_verify = preceding_phase_reports(
                    work,
                    state,
                    results,
                    state["current_phase"],
                )
                if state.get("status") == "closed":
                    phases_to_verify.append(
                        gate_snapshot(work, state, results, "closed", include_approvals=True)
                    )
                for report in phases_to_verify:
                    if not report["passed"]:
                        codes = ",".join(blocker["code"] for blocker in report["blockers"][:5])
                        failures.append(
                            f"{state['id']}: preceding gate invalid: {report['phase']} ({codes})"
                        )
            except GovernanceError as exc:
                failures.append(str(exc))
    for proposal in load_improvements():
        if proposal.get("status") not in {"pending", "approved", "applied", "rejected"}:
            failures.append(f"{proposal.get('id')}: invalid improvement status")
    if failures:
        for failure in failures:
            print(f"FAIL: {failure}")
        return 1
    for warning in warnings:
        print(f"WARN: {warning}")
    print("audit OK")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(dest="command", required=True)

    catalog = sub.add_parser("catalog", help="export or verify the machine-readable checklist")
    catalog.add_argument("--check", action="store_true")
    catalog.set_defaults(func=cmd_catalog)

    init = sub.add_parser("init", help="create a governed work item")
    init.add_argument("--id")
    init.add_argument("--title", required=True)
    init.add_argument("--request", required=True)
    init.add_argument("--profile", action="append", default=["CORE"])
    init.add_argument("--scope-file", help="scopeflow estimate used for task-specific checklist selection")
    init.add_argument("--actor", required=True)
    init.set_defaults(func=cmd_init)

    status = sub.add_parser("status", help="show current work item and gate status")
    status.add_argument("--work-item", required=True)
    status.add_argument("--json", action="store_true")
    status.set_defaults(func=cmd_status)

    inspect = sub.add_parser("inspect", help="strictly inspect a phase gate")
    inspect.add_argument("--work-item", required=True)
    inspect.add_argument("--phase")
    inspect.add_argument("--ignore-approvals", action="store_true")
    inspect.add_argument("--json", action="store_true")
    inspect.set_defaults(func=cmd_inspect)

    check = sub.add_parser("set-check", help="record one checklist result")
    check.add_argument("--work-item", required=True)
    check.add_argument("--item", required=True)
    check.add_argument("--applicability", required=True)
    check.add_argument("--verdict", required=True)
    check.add_argument("--severity", required=True)
    check.add_argument("--severity-rationale")
    check.add_argument("--reviewer", required=True)
    check.add_argument("--evidence", action="append")
    check.add_argument("--na-rationale")
    check.add_argument("--issue")
    check.add_argument("--remediation-plan")
    check.add_argument("--due-at")
    check.add_argument("--recheck-evidence", action="append")
    check.add_argument("--recheck-reviewer")
    check.add_argument("--exception-approver")
    check.add_argument("--exception-role")
    check.add_argument("--exception-rationale")
    check.add_argument("--exception-expires")
    check.add_argument("--clear-exception", action="store_true")
    check.set_defaults(func=cmd_set_check)

    checks = sub.add_parser("set-checks", help="atomically record checklist results from JSON")
    checks.add_argument("--work-item", required=True)
    checks.add_argument("--input", required=True)
    checks.add_argument("--actor", required=True)
    checks.set_defaults(func=cmd_set_checks)

    migrate = sub.add_parser("migrate", help="migrate a legacy work item to the current workflow schema")
    migrate.add_argument("--work-item", required=True)
    migrate.add_argument("--actor", required=True)
    migrate.set_defaults(func=cmd_migrate)

    approve = sub.add_parser("approve", help="deprecated alias for initial authorization")
    approve.add_argument("--work-item", required=True)
    approve.add_argument("--phase")
    approve.add_argument("--decision", choices=["approved", "rejected"], required=True)
    approve.add_argument("--approver", required=True)
    approve.add_argument("--role")
    approve.add_argument("--comment", required=True)
    approve.set_defaults(func=cmd_approve)

    authorize = sub.add_parser("authorize", help="record the single initial autonomous-execution decision")
    authorize.add_argument("--work-item", required=True)
    authorize.add_argument("--decision", choices=["approved", "rejected"], required=True)
    authorize.add_argument("--approver", required=True)
    authorize.add_argument("--comment", required=True)
    authorize.set_defaults(func=cmd_approve, phase=None, role=None)

    advance = sub.add_parser("advance", help="advance after the current gate passes")
    advance.add_argument("--work-item", required=True)
    advance.add_argument("--actor", required=True)
    advance.set_defaults(func=cmd_advance)

    retro = sub.add_parser("session-retrospective", help="generate a deterministic session retrospective")
    retro.add_argument("--session-id", required=True)
    retro.add_argument("--cwd", default=str(ROOT))
    retro.set_defaults(func=cmd_session_retrospective)

    improve_list = sub.add_parser("improvement-list", help="list skill improvement proposals")
    improve_list.add_argument("--status", choices=["pending", "approved", "applied", "rejected"])
    improve_list.add_argument("--json", action="store_true")
    improve_list.set_defaults(func=cmd_improvement_list)

    audit = sub.add_parser("audit", help="audit catalog and tamper-evident records")
    audit.set_defaults(func=cmd_audit)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        return int(args.func(args))
    except GovernanceError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
