from __future__ import annotations

import unittest
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import quote_sheetname

import update_checklist

ROOT = Path(__file__).resolve().parents[1]


class ChecklistTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.workbook = load_workbook(ROOT / "checklist.xlsx", data_only=False)

    @classmethod
    def tearDownClass(cls) -> None:
        cls.workbook.close()

    def test_auditable_execution_columns_are_present_on_every_review_sheet(self) -> None:
        for sheet in update_checklist.REVIEW_SHEETS:
            ws = self.workbook[sheet]
            self.assertEqual(ws.max_column, 23)
            self.assertEqual(
                [ws.cell(1, column).value for column in range(9, 24)],
                update_checklist.EXT_HEADERS,
            )

    def test_summary_formulas_use_only_each_sheet_data_range(self) -> None:
        summary = self.workbook["サマリ"]
        for row, sheet in enumerate(update_checklist.REVIEW_SHEETS, 5):
            end_row = self.workbook[sheet].max_row
            quoted = quote_sheetname(sheet)
            self.assertEqual(summary.cell(row, 4).value, f"=COUNTA({quoted}!$A$2:$A${end_row})")
            for column, severity in zip(range(5, 9), ["Critical", "High", "Medium", "Low"]):
                self.assertEqual(
                    summary.cell(row, column).value,
                    f'=COUNTIF({quoted}!$D$2:$D${end_row},"{severity}")',
                )
        formulas = [
            cell.value
            for row in summary.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ]
        self.assertTrue(formulas)
        self.assertFalse(any("1048576" in formula for formula in formulas))

    def test_swebok_v4a_source_and_all_knowledge_areas_are_auditable(self) -> None:
        sources = self.workbook["出典マスター"]
        self.assertEqual(sources.cell(3, 8).value, "前版との差分")
        rows = {sources.cell(row, 1).value: row for row in range(4, sources.max_row + 1)}
        row = rows["SRC-SWEBOK-V4A"]
        self.assertIn("v4.0a", sources.cell(row, 2).value)
        self.assertIn(update_checklist.EXPECTED_SWEBOK_SHA256, sources.cell(row, 9).value)
        self.assertEqual(self.workbook["KA対応表"].max_row - 3, 18)

    def test_known_composite_controls_remain_split_into_atomic_checks(self) -> None:
        ai = {self.workbook["19_AI"].cell(row, 1).value for row in range(2, self.workbook["19_AI"].max_row + 1)}
        self.assertTrue(
            {
                "AI-004", "AI-052", "AI-053",
                "AI-011", "AI-054", "AI-055",
                "AI-014", "AI-056", "AI-057", "AI-058", "AI-059",
                "AI-030", "AI-060", "AI-061",
                "AI-043", "AI-062", "AI-063", "AI-064", "AI-065", "AI-066",
            }
            <= ai
        )
        azure = {self.workbook["17_Azure"].cell(row, 1).value for row in range(2, self.workbook["17_Azure"].max_row + 1)}
        self.assertTrue({"AZR-251", "AZR-258", "AZR-286"} <= azure)

    def test_additional_items_and_contextual_controls_are_preserved(self) -> None:
        self.assertEqual(self.workbook["追加項目管理"].max_row - 3, 65)
        cloud = self.workbook["14_クラウド"]
        by_id = {cloud.cell(row, 1).value: row for row in range(2, cloud.max_row + 1)}
        self.assertIn("SLO", cloud.cell(by_id["CLD-001"], 5).value)
        self.assertIn("RTO/RPO", cloud.cell(by_id["CLD-004"], 5).value)
        self.assertIn("ステートレス化", cloud.cell(by_id["CLD-011"], 6).value)


if __name__ == "__main__":
    unittest.main()
